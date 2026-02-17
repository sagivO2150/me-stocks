#!/usr/bin/env python3
"""
Analyze FTAI stock for rise events between 04/06/2024 and 21/11/2024.
A rise event is defined as a continuous period where the stock price is increasing,
from a local minimum to a local maximum.
"""

import yfinance as yf
import pandas as pd
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import json


def load_ftai_insider_trades():
    """Load FTAI insider trades from expanded_insider_trades.json."""
    try:
        with open('output CSVs/expanded_insider_trades.json', 'r') as f:
            data = json.load(f)
        
        # Find BSAI entries
        for stock in data.get('data', []):
            if stock.get('ticker') == 'BSAI':
                # Extract trade dates and convert to DD/MM/YYYY format
                trades = []
                for trade in stock.get('trades', []):
                    trade_date = trade.get('trade_date', '')
                    if trade_date:
                        # Convert from YYYY-MM-DD to DD/MM/YYYY
                        try:
                            dt = datetime.strptime(trade_date, '%Y-%m-%d')
                            formatted_date = dt.strftime('%d/%m/%Y')
                            trades.append({
                                'date': formatted_date,
                                'insider_name': trade.get('insider_name', ''),
                                'value': trade.get('value', '')
                            })
                        except:
                            continue
                return trades
        return []
    except Exception as e:
        print(f"Warning: Could not load insider trades: {e}")
        return []


def find_insider_purchases_in_range(start_date: str, end_date: str, insider_trades: List[Dict]) -> str:
    """Find insider purchase dates that fall within the given date range.
    
    Args:
        start_date: Start date in DD/MM/YYYY format
        end_date: End date in DD/MM/YYYY format
        insider_trades: List of insider trade dictionaries with 'date' key
        
    Returns:
        Comma-separated string of purchase dates, or empty string if none found
    """
    try:
        # Parse dates
        start_dt = datetime.strptime(start_date, '%d/%m/%Y')
        end_dt = datetime.strptime(end_date, '%d/%m/%Y')
        
        # Find matching trades
        matching_dates = []
        for trade in insider_trades:
            trade_dt = datetime.strptime(trade['date'], '%d/%m/%Y')
            if start_dt <= trade_dt <= end_dt:
                matching_dates.append(trade['date'])
        
        # Return unique dates, comma-separated
        return ', '.join(sorted(set(matching_dates)))
    except Exception as e:
        return ''


def identify_rise_events(df: pd.DataFrame, min_days: int = 4, min_growth_pct: float = 2.0, min_decline_pct: float = 1.5, min_recovery_pct: float = 2.0) -> List[Dict]:
    """
    Identify rise events in stock price data.
    
    A rise event starts at a local minimum and ends at a local maximum.
    A rise event only ends when there are 2 CONSECUTIVE days of MEANINGFUL decline.
    Small recoveries after declines don't reset the counter - only strong recoveries do.
    
    Args:
        df: DataFrame with Date index and Close prices
        min_days: Minimum number of trading days for a valid rise event (default: 4)
        min_growth_pct: Minimum growth percentage for a valid rise event (default: 2.0%)
        min_decline_pct: Minimum decline percentage per day to count as meaningful dip (default: 1.5%)
        min_recovery_pct: Minimum recovery percentage to reset consecutive dips counter (default: 2.0%)
        
    Returns:
        List of dictionaries containing start_date, end_date, and growth_pct
    """
    rise_events = []
    
    if len(df) < 3:
        return rise_events
    
    # Track the current rise event
    in_rise_event = False
    current_start_idx = None
    current_start_price = None
    peak_idx = None
    peak_price = None
    consecutive_dips = 0
    
    # Track bottom hunting - finding the local minimum after a decline
    hunting_bottom = True
    potential_bottom_idx = 0
    potential_bottom_price = df['Close'].iloc[0]
    
    for i in range(len(df)):
        current_price = df['Close'].iloc[i]
        
        if hunting_bottom:
            # We're looking for a local minimum (bottom)
            # Keep tracking the lowest price
            if current_price <= potential_bottom_price:
                potential_bottom_idx = i
                potential_bottom_price = current_price
            
            # If price starts going up, we found the bottom
            if i > 0 and current_price > df['Close'].iloc[i - 1]:
                # Start a rise event from the bottom
                in_rise_event = True
                hunting_bottom = False
                current_start_idx = potential_bottom_idx
                current_start_price = potential_bottom_price
                peak_idx = i
                peak_price = current_price
                consecutive_dips = 0
            continue
        
        # We're in a rise event - check if price went up or down compared to previous day
        prev_price = df['Close'].iloc[i - 1]
        
        if current_price > prev_price:
            # Price went up - check if it's a meaningful recovery
            recovery_pct = ((current_price - prev_price) / prev_price) * 100
            
            if recovery_pct >= min_recovery_pct:
                # Meaningful recovery - reset consecutive dips counter
                consecutive_dips = 0
            # else: Small recovery after decline - keep counter as is
            
            # Update peak if this is a new high
            if current_price > peak_price:
                peak_idx = i
                peak_price = current_price
                consecutive_dips = 0  # New peak always resets counter
                
        elif current_price < prev_price:
            # Price went down - check if it's a meaningful decline
            decline_pct = ((prev_price - current_price) / prev_price) * 100
            
            if decline_pct >= min_decline_pct:
                # Meaningful decline - increment counter
                consecutive_dips += 1
            # else: Minor dip/plateau - keep counter as is (don't increment, don't reset)
            
            # If we have 2 consecutive dips, end the rise event
            if consecutive_dips >= 2 and peak_idx > current_start_idx:
                start_date = df.index[current_start_idx]
                end_date = df.index[peak_idx]
                growth_pct = ((peak_price - current_start_price) / current_start_price) * 100
                # Count trading days, not calendar days
                days_duration = peak_idx - current_start_idx + 1
                
                # Only add if meets minimum thresholds
                if days_duration >= min_days and growth_pct >= min_growth_pct:
                    rise_events.append({
                        'start_date': start_date.strftime('%d/%m/%Y'),
                        'end_date': end_date.strftime('%d/%m/%Y'),
                        'days': days_duration,
                        'growth_pct': round(growth_pct, 2)
                    })
                
                # Start hunting for the next bottom
                in_rise_event = False
                hunting_bottom = True
                potential_bottom_idx = i
                potential_bottom_price = current_price
                consecutive_dips = 0
        else:
            # Price unchanged - don't increment consecutive dips
            pass
    
    # Handle the last rise event if it's still active at the end
    if in_rise_event and peak_idx is not None and peak_idx > current_start_idx:
        start_date = df.index[current_start_idx]
        end_date = df.index[peak_idx]
        growth_pct = ((peak_price - current_start_price) / current_start_price) * 100
        # Count trading days, not calendar days
        days_duration = peak_idx - current_start_idx + 1
        
        # Only add if meets minimum thresholds
        if days_duration >= min_days and growth_pct >= min_growth_pct:
            rise_events.append({
                'start_date': start_date.strftime('%d/%m/%Y'),
                'end_date': end_date.strftime('%d/%m/%Y'),
                'days': days_duration,
                'growth_pct': round(growth_pct, 2)
            })
    
    return rise_events


def main():
    """Main function to analyze FTAI stock and identify rise events."""
    ticker = "BSAI"
    start_date = "2015-10-05"
    end_date = "2026-02-12"
    
    print(f"Fetching {ticker} data from {start_date} to {end_date}...")
    
    try:
        # Load insider trades
        insider_trades = load_ftai_insider_trades()
        print(f"Loaded {len(insider_trades)} insider trades for {ticker}")
        print()
        
        # Fetch stock data
        stock = yf.Ticker(ticker)
        df = stock.history(start=start_date, end=end_date)
        
        if df.empty:
            print(f"No data found for {ticker} in the specified date range.")
            return
        
        print(f"Retrieved {len(df)} trading days of data")
        print(f"Date range: {df.index[0].strftime('%Y-%m-%d')} to {df.index[-1].strftime('%Y-%m-%d')}")
        print(f"Price range: ${df['Close'].min():.2f} to ${df['Close'].max():.2f}")
        print()
        
        # Identify rise events
        rise_events = identify_rise_events(df)
        
        print(f"Found {len(rise_events)} rise events:")
        print()
        
        # Create combined list with rise events and down periods
        combined_events = []
        cumulative_multiplier = 1.0  # Track actual price multiplier from start
        
        for i, rise_event in enumerate(rise_events):
            # Add the rise event
            cumulative_multiplier *= (1 + rise_event['growth_pct'] / 100)
            cumulative_pct = (cumulative_multiplier - 1) * 100
            
            insider_purchases = find_insider_purchases_in_range(
                rise_event['start_date'], 
                rise_event['end_date'], 
                insider_trades
            )
            combined_events.append({
                'event_type': 'RISE',
                'start_date': rise_event['start_date'],
                'end_date': rise_event['end_date'],
                'days': rise_event['days'],
                'change_pct': rise_event['growth_pct'],
                'cumulative_pct': round(cumulative_pct, 2),
                'insider_purchases': insider_purchases
            })
            
            # Calculate down period to next rise (if not the last rise event)
            if i < len(rise_events) - 1:
                # Find the dates in the dataframe
                current_end_str = rise_event['end_date']
                next_start_str = rise_events[i + 1]['start_date']
                
                # Get the peak price at end of current rise
                peak_mask = df.index.strftime('%d/%m/%Y') == current_end_str
                peak_price = df.loc[peak_mask, 'Close'].iloc[0]
                
                # Get the bottom price at start of next rise
                bottom_mask = df.index.strftime('%d/%m/%Y') == next_start_str
                bottom_price = df.loc[bottom_mask, 'Close'].iloc[0]
                
                # Calculate decline percentage
                decline_pct = ((peak_price - bottom_price) / peak_price) * 100
                
                # Calculate number of trading days
                peak_idx = df.index[peak_mask][0]
                bottom_idx = df.index[bottom_mask][0]
                down_period_df = df.loc[peak_idx:bottom_idx]
                down_days = len(down_period_df) - 1  # Exclude the start date since it's the peak
                
                if down_days > 0:
                    cumulative_multiplier *= (1 - decline_pct / 100)
                    cumulative_pct = (cumulative_multiplier - 1) * 100
                    
                    insider_purchases = find_insider_purchases_in_range(
                        rise_event['end_date'],
                        rise_events[i + 1]['start_date'],
                        insider_trades
                    )
                    combined_events.append({
                        'event_type': 'DOWN',
                        'start_date': rise_event['end_date'],
                        'end_date': rise_events[i + 1]['start_date'],
                        'days': down_days,
                        'change_pct': -round(decline_pct, 2),
                        'cumulative_pct': round(cumulative_pct, 2),
                        'insider_purchases': insider_purchases
                    })
        
        # Create DataFrame from combined events
        results_df = pd.DataFrame(combined_events)
        
        # Calculate ranks within each event type based on total percentage
        # For RISE events: lower % = worse rank (1 is worst)
        # For DOWN events: larger decline % = worse rank (1 is worst)
        rise_events_df = results_df[results_df['event_type'] == 'RISE'].copy()
        down_events_df = results_df[results_df['event_type'] == 'DOWN'].copy()
        
        rise_count = len(rise_events_df)
        down_count = len(down_events_df)
        
        if not rise_events_df.empty:
            # For rises, ascending=False means higher % gets rank 1 (best), lower % gets higher rank (worst)
            rise_events_df['rank'] = rise_events_df['change_pct'].rank(ascending=False, method='min').astype(int)
            rise_events_df['rank_display'] = rise_events_df['rank'].apply(lambda x: f"{x}/{rise_count}")
        
        if not down_events_df.empty:
            # For downs, ascending=False means less negative gets rank 1 (best), more negative gets higher rank (worst)
            down_events_df['rank'] = down_events_df['change_pct'].rank(ascending=False, method='min').astype(int)
            down_events_df['rank_display'] = down_events_df['rank'].apply(lambda x: f"{x}/{down_count}")
        
        # Merge back into combined_events
        for i, event in enumerate(combined_events):
            if event['event_type'] == 'RISE':
                idx = rise_events_df[
                    (rise_events_df['start_date'] == event['start_date']) & 
                    (rise_events_df['end_date'] == event['end_date'])
                ].index[0]
                event['rank'] = rise_events_df.loc[idx, 'rank_display']
            else:
                idx = down_events_df[
                    (down_events_df['start_date'] == event['start_date']) & 
                    (down_events_df['end_date'] == event['end_date'])
                ].index[0]
                event['rank'] = down_events_df.loc[idx, 'rank_display']
        
        # Recreate DataFrame with ranks
        results_df = pd.DataFrame(combined_events)
        
        if not results_df.empty:
            # Print results
            for idx, event in enumerate(combined_events, 1):
                if event['event_type'] == 'RISE':
                    print(f"{idx}. RISE: {event['start_date']} → {event['end_date']} ({event['days']} days): +{event['change_pct']}% | Rank: {event['rank']} | Cumulative: {event['cumulative_pct']}%")
                else:
                    print(f"{idx}. DOWN: {event['start_date']} → {event['end_date']} ({event['days']} days): {event['change_pct']}% | Rank: {event['rank']} | Cumulative: {event['cumulative_pct']}%")
            
            # Save to CSV
            output_file = "output CSVs/ftai_rise_events.csv"
            results_df.to_csv(output_file, index=False)
            print()
            print(f"Results saved to: {output_file}")
            
            # Save to Excel with colors
            excel_file = "output CSVs/ftai_rise_events.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "FTAI Rise Events"
            
            # Define colors
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            bold_font = Font(bold=True)
            
            # Write headers
            headers = ['Event Type', 'Start Date', 'End Date', 'Days', 'Change %', 'Rank', 'Cumulative %', 'Insider Purchases']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = bold_font
            
            # Write data with colors
            for row_idx, event in enumerate(combined_events, 2):
                ws.cell(row=row_idx, column=1, value=event['event_type'])
                ws.cell(row=row_idx, column=2, value=event['start_date'])
                ws.cell(row=row_idx, column=3, value=event['end_date'])
                ws.cell(row=row_idx, column=4, value=event['days'])
                ws.cell(row=row_idx, column=5, value=event['change_pct'])
                ws.cell(row=row_idx, column=6, value=event['rank'])
                ws.cell(row=row_idx, column=7, value=event['cumulative_pct'])
                ws.cell(row=row_idx, column=8, value=event.get('insider_purchases', ''))
                
                # Apply color to entire row
                fill = green_fill if event['event_type'] == 'RISE' else red_fill
                for col_idx in range(1, 9):
                    ws.cell(row=row_idx, column=col_idx).fill = fill
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 8
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 8
            ws.column_dimensions['G'].width = 14
            ws.column_dimensions['H'].width = 20
            
            wb.save(excel_file)
            print(f"Colored Excel file saved to: {excel_file}")
            
            # Summary statistics for rise events only
            rise_only = [e for e in combined_events if e['event_type'] == 'RISE']
            total_growth = sum([e['change_pct'] for e in rise_only])
            avg_growth = total_growth / len(rise_only) if rise_only else 0
            max_growth = max([e['change_pct'] for e in rise_only]) if rise_only else 0
            
            print()
            print("Summary:")
            print(f"Total growth across all rise events: {total_growth:.2f}%")
            print(f"Average growth per rise event: {avg_growth:.2f}%")
            print(f"Maximum single rise event: {max_growth:.2f}%")
        else:
            print("No rise events found in the specified period.")
    
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
