#!/usr/bin/env python3
"""
Generate detailed analysis files (CSV/XLSX/JSON) for a specific stock.
This script provides drill-down visibility into rise/fall events, mid-rises, mid-falls, etc.

Usage:
    .venv/bin/python /path/to/generate_stock_detailed_analysis.py TICKER

Example:
    .venv/bin/python generate_stock_detailed_analysis.py GROV

Outputs:
    - output CSVs/{ticker}_rise_events.csv
    - output CSVs/{ticker}_rise_events.xlsx (color-coded)
    - output CSVs/{ticker}_rise_volatility_analysis.json
"""

import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
import json
from typing import List, Dict, Optional, Tuple
from enum import Enum
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


class MarketPhase(Enum):
    """Current market phase we're observing."""
    RISING = "rising"
    FALLING = "falling"
    UNKNOWN = "unknown"


class TradingState:
    """Track the current state of our live trading simulation."""
    
    def __init__(self):
        # Market observation
        self.phase = MarketPhase.UNKNOWN
        self.last_peak_date = None
        
        # Price history for rise start detection
        self.price_history = []
        
        # Dip-recovery-dip tracking for fall detection
        self.first_dip_date = None
        self.first_dip_price = None
        self.dip_low_price = None
        self.in_recovery = False
        self.recovery_high = None
        self.consecutive_up_days = 0
        
        # Current trend tracking
        self.trend_start_date = None
        self.trend_start_price = None
        self.trend_peak_price = None
        self.trend_peak_date = None
        self.trend_low_price = float('inf')
        self.trend_low_date = None
        
        # Event tracking for CSV/Excel output
        self.all_events = []
        
        # Previous trend memory
        self.prev_fall_pct = 0
        self.prev_rise_pct = 0
        self.prev_rise_start_price = None
        self.prev_rise_peak_price = None
        self.prev_fall_start_price = None
        
        # Insider activity tracking
        self.insiders_bought_in_rise = []
        self.insiders_bought_in_fall = []
        self.shopping_spree_peak_price = None
        
        # Position tracking
        self.in_position = False
        self.entry_date = None
        self.entry_price = None
        self.target_price = None
        self.buy_type = None
        self.position_size = 2000
        self.max_mid_fall_before_target = 0
        self.target_reached = False
        self.peak_since_entry = 0
        
        # Absorption buy specific tracking
        self.rise_start_price = None
        self.in_mid_rise = False
    
    def update_phase(self, current_price: float, prev_price: float, date: datetime):
        """Update market phase using FIRST DIP → RECOVERY → SECOND DIP pattern for fall detection."""
        # Track price history
        self.price_history.append((date, current_price))
        if len(self.price_history) > 4:
            self.price_history.pop(0)
        
        # Treat ±$0.01 as plateau
        price_diff = abs(current_price - prev_price)
        is_plateau = price_diff <= 0.01
        is_up = current_price > prev_price and not is_plateau
        is_down = current_price < prev_price and not is_plateau
        
        # Track consecutive up days
        if is_up:
            self.consecutive_up_days += 1
        else:
            self.consecutive_up_days = 0
        
        if self.phase == MarketPhase.UNKNOWN or self.phase == MarketPhase.FALLING:
            # HUNTING FOR RISE START: Need 2 consecutive up days (or 3 for hot stocks)
            is_hot_stock = bool(self.insiders_bought_in_rise and self.insiders_bought_in_fall)
            required_up_days = 3 if is_hot_stock else 2
            
            if self.consecutive_up_days >= required_up_days:
                days_since_last_peak = 999
                
                if days_since_last_peak >= 0:
                    # Clear old insider data if no recent activity (30+ days)
                    if not self.in_position:
                        days_since_last_purchase = 999
                        for purchase_list in [self.insiders_bought_in_rise, self.insiders_bought_in_fall]:
                            for purchase in purchase_list:
                                purchase_date = pd.to_datetime(purchase['date']).tz_localize(None)
                                current_date_naive = date.tz_localize(None) if hasattr(date, 'tz_localize') else date
                                days_diff = (current_date_naive - purchase_date).days
                                days_since_last_purchase = min(days_since_last_purchase, days_diff)
                        
                        if days_since_last_purchase > 30:
                            self.insiders_bought_in_rise = []
                            self.insiders_bought_in_fall = []
                            self.shopping_spree_peak_price = None
                    
                    # Find the actual rise start
                    lookback = 4 if is_hot_stock else 3
                    
                    if len(self.price_history) >= lookback:
                        bottom_date, bottom_price = self.price_history[-lookback]
                        actual_start_date = bottom_date
                        actual_start_price = bottom_price
                    else:
                        actual_start_date = date
                        actual_start_price = prev_price
                    
                    # Record the completed FALL event if we were falling before
                    if self.phase == MarketPhase.FALLING and self.trend_start_date:
                        fall_days = (actual_start_date - self.trend_start_date).days
                        fall_pct = ((self.trend_low_price - self.trend_start_price) / self.trend_start_price) * 100
                        
                        if not self.in_position:
                            self.prev_fall_pct = abs(fall_pct)
                            self.prev_fall_start_price = self.trend_start_price
                        
                        # Filter insiders who bought during THIS fall only
                        fall_start_naive = self.trend_start_date.tz_localize(None) if hasattr(self.trend_start_date, 'tz_localize') else self.trend_start_date
                        fall_end_naive = actual_start_date.tz_localize(None) if hasattr(actual_start_date, 'tz_localize') else actual_start_date
                        
                        fall_insiders = [
                            i for i in self.insiders_bought_in_fall 
                            if fall_start_naive <= pd.to_datetime(i['date']).tz_localize(None) <= fall_end_naive
                        ]
                        
                        self.all_events.append({
                            'event_type': 'DOWN',
                            'start_date': self.trend_start_date,
                            'end_date': actual_start_date,
                            'days': fall_days,
                            'change_pct': fall_pct,
                            'start_price': self.trend_start_price,
                            'end_price': self.trend_low_price,
                            'insiders': fall_insiders
                        })
                    
                    # Start rise
                    self.phase = MarketPhase.RISING
                    self.trend_start_date = actual_start_date
                    self.trend_start_price = actual_start_price
                    self.trend_peak_price = current_price
                    self.trend_peak_date = date
                    self.first_dip_date = None
                    self.first_dip_price = None
                    self.in_recovery = False
        
        elif self.phase == MarketPhase.RISING:
            # Update peak if new high
            if current_price > self.trend_peak_price:
                self.trend_peak_price = current_price
                self.trend_peak_date = date
                self.first_dip_date = None
                self.first_dip_price = None
                self.dip_low_price = None
                self.in_recovery = False
            
            # FALL DETECTION: First dip → recovery → second dip
            if is_down:
                decline_from_peak = ((self.trend_peak_price - current_price) / self.trend_peak_price) * 100
                
                if self.first_dip_date is None:
                    if decline_from_peak >= 1.0:
                        self.first_dip_date = date
                        self.first_dip_price = current_price
                        self.dip_low_price = current_price
                        self.in_recovery = False
                elif not self.in_recovery:
                    if current_price < self.dip_low_price:
                        self.dip_low_price = current_price
                elif self.in_recovery:
                    decline_from_recovery = ((self.recovery_high - current_price) / self.recovery_high) * 100
                    if decline_from_recovery >= 1.0:
                        self.prev_rise_pct = ((self.trend_peak_price - self.trend_start_price) / 
                                             self.trend_start_price) * 100
                        self.prev_rise_start_price = self.trend_start_price
                        self.prev_rise_peak_price = self.trend_peak_price
                        self.last_peak_date = self.trend_peak_date
                        
                        # Transition to FALLING
                        self.phase = MarketPhase.FALLING
                        
                        from pandas.tseries.offsets import BDay
                        actual_rise_end = self.first_dip_date - BDay(1)
                        
                        # Record the completed RISE event
                        rise_days = (actual_rise_end - self.trend_start_date).days
                        rise_pct = self.prev_rise_pct
                        
                        rise_start_naive = self.trend_start_date.tz_localize(None) if hasattr(self.trend_start_date, 'tz_localize') else self.trend_start_date
                        rise_end_naive = actual_rise_end.tz_localize(None) if hasattr(actual_rise_end, 'tz_localize') else actual_rise_end
                        
                        rise_insiders = [
                            i for i in self.insiders_bought_in_rise 
                            if rise_start_naive <= pd.to_datetime(i['date']).tz_localize(None) <= rise_end_naive
                        ]
                        
                        self.all_events.append({
                            'event_type': 'RISE',
                            'start_date': self.trend_start_date,
                            'end_date': actual_rise_end,
                            'days': rise_days,
                            'change_pct': rise_pct,
                            'start_price': self.trend_start_price,
                            'end_price': self.trend_peak_price,
                            'insiders': rise_insiders
                        })
                        
                        self.trend_start_date = self.first_dip_date
                        self.trend_start_price = self.trend_peak_price
                        self.trend_low_price = current_price
                        self.trend_low_date = date
                        self.consecutive_up_days = 0
                        
                        self.first_dip_date = None
                        self.first_dip_price = None
                        self.dip_low_price = None
                        self.in_recovery = False
            
            elif is_up and self.in_recovery:
                self.first_dip_date = None
                self.first_dip_price = None
                self.dip_low_price = None
                self.in_recovery = False
                self.recovery_high = current_price
            
            elif (is_up or is_plateau) and self.first_dip_date is not None and not self.in_recovery:
                recovery_from_dip = ((current_price - self.dip_low_price) / self.dip_low_price) * 100
                if recovery_from_dip >= 0.0:
                    self.in_recovery = True
                    self.recovery_high = current_price
        
        # Track lowest price during fall
        if self.phase == MarketPhase.FALLING:
            if current_price < self.trend_low_price:
                self.trend_low_price = current_price
                self.trend_low_date = date
        
        # Calculate current fall percentage
        if self.phase == MarketPhase.FALLING and self.trend_start_price and not self.in_position:
            self.prev_fall_pct = ((self.trend_start_price - current_price) / self.trend_start_price) * 100
    
    def record_insider_purchase(self, date_str: str, trade_info: Dict):
        """Record an insider purchase occurring today."""
        trade_data = {
            'date': date_str,
            'price': trade_info['price'],
            'insider_name': trade_info['insider_name'],
            'value': trade_info['value'],
            'stock_price': trade_info.get('stock_price', trade_info['price'])
        }
        
        if self.phase == MarketPhase.RISING:
            self.insiders_bought_in_rise.append(trade_data)
            if self.shopping_spree_peak_price is None or trade_data['stock_price'] > self.shopping_spree_peak_price:
                self.shopping_spree_peak_price = trade_data['stock_price']
        elif self.phase == MarketPhase.FALLING:
            self.insiders_bought_in_fall.append(trade_data)
            if self.shopping_spree_peak_price is None or trade_data['stock_price'] > self.shopping_spree_peak_price:
                self.shopping_spree_peak_price = trade_data['stock_price']


def analyze_rise_volatility(df: pd.DataFrame, rise_event: Dict) -> Dict:
    """Analyze the volatility pattern during and after a rise event."""
    start_date = rise_event['start_date']
    end_date = rise_event['end_date']
    
    rise_df = df.loc[start_date:end_date]
    
    try:
        end_idx = df.index.get_indexer([end_date], method='nearest')[0]
    except:
        return None
        
    post_rise_end_idx = min(end_idx + 31, len(df))
    post_rise_df = df.iloc[end_idx:post_rise_end_idx]
    
    result = {
        'rise_start_date': start_date.strftime('%d/%m/%Y'),
        'rise_end_date': end_date.strftime('%d/%m/%Y'),
        'rise_days': rise_event['days'],
        'rise_percentage': rise_event['change_pct'],
        'mid_rises': {},
        'mid_falls': {},
        'first_dip': None,
        'first_recovery': None,
        'second_dip': None,
        'insiders': rise_event.get('insiders', [])
    }
    
    # Track consecutive movements during the rise
    if len(rise_df) > 1:
        i = 0
        while i < len(rise_df) - 1:
            movement_start_price = rise_df['Close'].iloc[i]
            movement_start_idx = i
            current_direction = None
            
            j = i + 1
            while j < len(rise_df):
                prev_price = rise_df['Close'].iloc[j-1]
                current_price = rise_df['Close'].iloc[j]
                
                if current_price > prev_price:
                    day_direction = 'up'
                elif current_price < prev_price:
                    day_direction = 'down'
                else:
                    j += 1
                    continue
                
                if current_direction is None:
                    current_direction = day_direction
                    j += 1
                    continue
                
                if day_direction != current_direction:
                    break
                    
                j += 1
            
            if current_direction is not None and j > i + 1:
                movement_end_price = rise_df['Close'].iloc[j-1]
                movement_end_date = rise_df.index[j-1]
                total_change_pct = ((movement_end_price - movement_start_price) / movement_start_price) * 100
                
                if current_direction == 'up' and total_change_pct >= 1.0:
                    pct_key = str(round(total_change_pct, 2))
                    result['mid_rises'][pct_key] = {
                        'date': movement_end_date.strftime('%d/%m/%Y')
                    }
                elif current_direction == 'down' and total_change_pct <= -1.0:
                    pct_key = str(round(total_change_pct, 2))
                    result['mid_falls'][pct_key] = {
                        'date': movement_end_date.strftime('%d/%m/%Y')
                    }
                
                i = j - 1
            else:
                i += 1
    
    # Analyze post-rise behavior
    if len(post_rise_df) > 1:
        peak_price = post_rise_df['Close'].iloc[0]
        
        first_dip_found = False
        first_dip_low = peak_price
        first_dip_low_idx = 0
        
        first_recovery_found = False
        recovery_high = peak_price
        recovery_high_idx = 0
        
        for i in range(1, len(post_rise_df)):
            current_price = post_rise_df['Close'].iloc[i]
            current_date = post_rise_df.index[i]
            
            if not first_dip_found:
                if current_price < first_dip_low:
                    first_dip_low = current_price
                    first_dip_low_idx = i
                    fall_pct = ((peak_price - current_price) / peak_price) * 100
                    if fall_pct >= 1.0:
                        result['first_dip'] = {
                            'date': current_date.strftime('%d/%m/%Y'),
                            'percentage': round(-fall_pct, 2),
                            'days_after_peak': i
                        }
                elif current_price > first_dip_low and result['first_dip'] is not None:
                    recovery_from_low = ((current_price - first_dip_low) / first_dip_low) * 100
                    if recovery_from_low >= 1.0:
                        first_dip_found = True
                        recovery_high = current_price
                        recovery_high_idx = i
                        result['first_recovery'] = {
                            'date': current_date.strftime('%d/%m/%Y'),
                            'percentage': round(recovery_from_low, 2),
                            'days_after_peak': i
                        }
                    
            elif not first_recovery_found:
                if current_price > recovery_high:
                    recovery_high = current_price
                    recovery_high_idx = i
                    recovery_pct = ((current_price - first_dip_low) / first_dip_low) * 100
                    if recovery_pct >= 1.0:
                        result['first_recovery'] = {
                            'date': current_date.strftime('%d/%m/%Y'),
                            'percentage': round(recovery_pct, 2),
                            'days_after_peak': i
                        }
                elif current_price < recovery_high and result['first_recovery'] is not None:
                    first_recovery_found = True
                    
            else:
                second_dip_pct = ((recovery_high - current_price) / recovery_high) * 100
                if second_dip_pct >= 1.0:
                    days_since_recovery = i - recovery_high_idx
                    if result['second_dip'] is None or second_dip_pct > abs(result['second_dip']['percentage']):
                        result['second_dip'] = {
                            'date': current_date.strftime('%d/%m/%Y'),
                            'percentage': round(-second_dip_pct, 2),
                            'days_after_peak': i,
                            'days_since_recovery': days_since_recovery
                        }
    
    return result


def generate_event_files(ticker: str, events: List[Dict], price_df: pd.DataFrame):
    """Generate CSV and Excel files for rise/fall events."""
    if not events:
        print("⚠️ No events to export")
        return
    
    # Calculate cumulative percentages and ranks
    cumulative_pct = 0
    rise_events = [e for e in events if e['event_type'] == 'RISE']
    fall_events = [e for e in events if e['event_type'] == 'DOWN']
    
    rise_events_sorted = sorted(rise_events, key=lambda x: abs(x['change_pct']), reverse=True)
    fall_events_sorted = sorted(fall_events, key=lambda x: abs(x['change_pct']), reverse=True)
    
    rise_ranks = {id(e): f"{i+1}/{len(rise_events)}" for i, e in enumerate(rise_events_sorted)}
    fall_ranks = {id(e): f"{i+1}/{len(fall_events)}" for i, e in enumerate(fall_events_sorted)}
    
    # Prepare data for export
    export_data = []
    for event in events:
        cumulative_pct += event['change_pct']
        
        insiders_str = ""
        if event['insiders']:
            insider_dates = sorted(set([i['date'] for i in event['insiders']]))
            insiders_str = ", ".join([datetime.strptime(d, '%Y-%m-%d').strftime('%d/%m/%Y') for d in insider_dates])
        
        rank = rise_ranks.get(id(event)) if event['event_type'] == 'RISE' else fall_ranks.get(id(event))
        
        export_data.append({
            'event_type': event['event_type'],
            'start_date': event['start_date'].strftime('%d/%m/%Y'),
            'end_date': event['end_date'].strftime('%d/%m/%Y'),
            'days': event['days'],
            'change_pct': round(event['change_pct'], 2),
            'cumulative_pct': round(cumulative_pct, 2),
            'insider_purchases': insiders_str if insiders_str else None,
            'rank': rank
        })
    
    # Save to CSV
    csv_file = f'output CSVs/{ticker.lower()}_rise_events.csv'
    df = pd.DataFrame(export_data)
    df.to_csv(csv_file, index=False)
    print(f"✓ CSV saved to: {csv_file}")
    
    # Create Excel with colors
    wb = Workbook()
    ws = wb.active
    ws.title = f"{ticker} Rise-Fall Events"
    
    # Colors
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    rise_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    fall_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    # Headers
    headers = ['Event Type', 'Start Date', 'End Date', 'Days', 'Change %', 'Rank', 'Cumulative %', 'Insider Purchases']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
    
    # Data rows
    for row_data in export_data:
        ws.append([
            row_data['event_type'],
            row_data['start_date'],
            row_data['end_date'],
            row_data['days'],
            row_data['change_pct'],
            row_data['rank'],
            row_data['cumulative_pct'],
            row_data['insider_purchases']
        ])
        
        fill = rise_fill if row_data['event_type'] == 'RISE' else fall_fill
        for cell in ws[ws.max_row]:
            cell.fill = fill
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 30
    
    # Save Excel
    excel_file = f'output CSVs/{ticker.lower()}_rise_events.xlsx'
    wb.save(excel_file)
    print(f"✓ Excel saved to: {excel_file}")


def generate_volatility_json(ticker: str, events: List[Dict], df: pd.DataFrame):
    """Generate JSON with volatility analysis for all rise events."""
    rise_events = [e for e in events if e['event_type'] == 'RISE']
    
    volatility_analysis = {}
    for rise_event in rise_events:
        analysis = analyze_rise_volatility(df, rise_event)
        
        if analysis is None:
            continue
        
        pct_key = str(round(rise_event['change_pct'], 2))
        volatility_analysis[pct_key] = analysis
    
    json_output = {
        'ticker': ticker,
        'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'total_rise_events': len(rise_events),
        'rise_events': volatility_analysis
    }
    
    json_file = f'output CSVs/{ticker.lower()}_rise_volatility_analysis.json'
    with open(json_file, 'w') as f:
        json.dump(json_output, f, indent=2)
    
    print(f"✓ Volatility JSON saved to: {json_file}")


def main():
    """Generate detailed analysis files for a specific stock."""
    if len(sys.argv) < 2:
        print("Usage: python generate_stock_detailed_analysis.py TICKER")
        print("Example: python generate_stock_detailed_analysis.py GROV")
        sys.exit(1)
    
    ticker = sys.argv[1].upper()
    
    print("=" * 80)
    print(f"GENERATING DETAILED ANALYSIS FOR {ticker}")
    print("=" * 80)
    print()
    
    # Load insider trades database
    print("Loading insider trades database...")
    with open('output CSVs/expanded_insider_trades.json', 'r') as f:
        data = json.load(f)
    
    # Find ticker in database
    stock_data = None
    for stock in data.get('data', []):
        if stock.get('ticker', '') == ticker:
            stock_data = stock
            break
    
    if not stock_data:
        print(f"❌ Ticker {ticker} not found in database")
        sys.exit(1)
    
    # Get insider trades for this ticker
    insider_trades = {}
    for trade in stock_data.get('trades', []):
        trade_date = trade.get('trade_date', '')
        if trade_date:
            try:
                price_str = str(trade.get('price', '0')).replace('$', '').replace(',', '')
                price = float(price_str) if price_str else 0.0
                
                value_str = str(trade.get('value', '0')).replace('$', '').replace('+', '').replace(',', '')
                value = float(value_str) if value_str else 0.0
                
                trade_info = {
                    'price': price,
                    'insider_name': trade.get('insider_name', ''),
                    'value': value,
                    'title': trade.get('title', '')
                }
                
                if trade_date not in insider_trades:
                    insider_trades[trade_date] = []
                insider_trades[trade_date].append(trade_info)
            except:
                continue
    
    print(f"✓ Loaded {len(insider_trades)} insider trade dates")
    
    # Fetch stock price data
    print(f"Fetching price data for {ticker}...")
    stock = yf.Ticker(ticker)
    price_df = stock.history(period="max")
    
    if price_df.empty:
        print(f"❌ No price data found for {ticker}")
        sys.exit(1)
    
    print(f"✓ Price data: {price_df.index[0].strftime('%Y-%m-%d')} to {price_df.index[-1].strftime('%Y-%m-%d')} ({len(price_df)} days)")
    print()
    
    # Run simulation to track events
    state = TradingState()
    
    for i in range(1, len(price_df)):
        current_date = price_df.index[i]
        current_price = price_df['Close'].iloc[i]
        prev_price = price_df['Close'].iloc[i-1]
        date_str = current_date.strftime('%Y-%m-%d')
        
        state.update_phase(current_price, prev_price, current_date)
        
        if date_str in insider_trades:
            for trade_info in insider_trades[date_str]:
                trade_info_with_price = trade_info.copy()
                trade_info_with_price['stock_price'] = current_price
                state.record_insider_purchase(date_str, trade_info_with_price)
    
    print("=" * 80)
    print("GENERATING OUTPUT FILES")
    print("=" * 80)
    
    # Generate all output files
    generate_event_files(ticker, state.all_events, price_df)
    generate_volatility_json(ticker, state.all_events, price_df)
    
    print()
    print("=" * 80)
    print(f"✓ Analysis complete for {ticker}")
    print("=" * 80)


if __name__ == "__main__":
    main()
