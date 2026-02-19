#!/usr/bin/env python3
"""
Backtest the Insider Conviction Trading Strategy for GROV.
NO HINDSIGHT - Simulates live trading day-by-day.

Strategy Rules (as if trading live):
1. BUY CONDITIONS (detected in real-time):
   a) Shopping Spree: Insiders buy during RISE, continue buying during the FALL
      - Previous decline must be ≥30%
      - Buy when we detect stock recovering (2+ days of rise after insider buying)
      - Target: Peak price during the shopping spree period
   
   b) Absorption Buy: Insiders buy ONLY during fall (not during rise)
      - Require minimum $5K total insider investment during the fall
      - Buy when we detect recovery starting (2+ consecutive up days)
      - Target: Recover the fall magnitude (cumulative mid-rises >= fall percentage)

2. SELL CONDITIONS (real-time detection):
   - Never sell during continuous rises
   - After target reached: Sell on first significant dip
   - Significant = larger than largest dip seen before hitting target
   - Emergency stop: -15% from entry

3. LIVE DETECTION LOGIC:
   - Track daily price movements
   - Detect when we transition from fall to rise
   - Track insider purchases as they occur
   - Make buy/sell decisions based only on past data
"""

import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
import json
from typing import List, Dict, Optional, Tuple
from enum import Enum


class MarketPhase(Enum):
    """Current market phase we're observing."""
    RISING = "rising"
    FALLING = "falling"
    UNKNOWN = "unknown"


def load_grov_data():
    """Load BSFC stock data and insider trades."""
    # Load insider trades
    with open('output CSVs/expanded_insider_trades.json', 'r') as f:
        data = json.load(f)
    
    insider_trades = {}  # date -> list of trade info
    for stock in data.get('data', []):
        if stock.get('ticker') == 'BSFC':
            for trade in stock.get('trades', []):
                trade_date = trade.get('trade_date', '')
                if trade_date:
                    try:
                        # Parse price (remove $ and commas)
                        price_str = str(trade.get('price', '0')).replace('$', '').replace(',', '')
                        price = float(price_str) if price_str else 0.0
                        
                        # Parse value (remove $, +, commas and convert to float)
                        value_str = str(trade.get('value', '0')).replace('$', '').replace('+', '').replace(',', '')
                        value = float(value_str) if value_str else 0.0
                        
                        trade_info = {
                            'price': price,
                            'insider_name': trade.get('insider_name', ''),
                            'value': value,  # Now numeric
                            'title': trade.get('title', '')
                        }
                        
                        # Store as list to handle multiple insiders on same date
                        if trade_date not in insider_trades:
                            insider_trades[trade_date] = []
                        insider_trades[trade_date].append(trade_info)
                    except:
                        continue
            break
    
    # Fetch stock price data
    stock = yf.Ticker("BSFC")
    price_df = stock.history(period="max")
    
    return insider_trades, price_df


class TradingState:
    """Track the current state of our live trading simulation."""
    
    def __init__(self):
        # Market observation
        self.phase = MarketPhase.UNKNOWN
        self.last_peak_date = None  # Track last peak to enforce 15-day gap before new rises
        
        # Price history for rise start detection (need 2 days back)
        self.price_history = []  # List of (date, price) tuples
        
        # Dip-recovery-dip tracking for fall detection
        self.first_dip_date = None
        self.first_dip_price = None
        self.dip_low_price = None  # Track lowest point during dip
        self.in_recovery = False
        self.recovery_high = None
        self.consecutive_up_days = 0  # Track consecutive up days for rise detection
        
        # Current trend tracking
        self.trend_start_date = None
        self.trend_start_price = None
        self.trend_peak_price = None
        self.trend_peak_date = None
        self.trend_low_price = float('inf')  # Start with very high value for bottom hunting
        self.trend_low_date = None
        
        # Event tracking for CSV/Excel output
        self.all_events = []  # List of all rise/fall events
        
        # Previous trend memory (for evaluating buy conditions)
        self.prev_fall_pct = 0
        self.prev_rise_pct = 0
        self.prev_rise_start_price = None
        self.prev_rise_peak_price = None
        self.prev_fall_start_price = None
        
        # Insider activity tracking
        self.insiders_bought_in_rise = []
        self.insiders_bought_in_fall = []
        self.shopping_spree_peak_price = None  # Track highest STOCK price during shopping spree
        
        # Position tracking
        self.in_position = False
        self.entry_date = None
        self.entry_price = None
        self.target_price = None
        self.buy_type = None
        self.position_size = 2000
        self.max_mid_fall_before_target = 0
        self.target_reached = False
        self.peak_since_entry = 0
        
        # Absorption buy specific tracking
        self.rise_start_price = None  # Track where the rise started for cumulative calculation
        self.cumulative_mid_rises_pct = 0  # Sum of all mid-rise percentages in current rise event
        self.mid_rise_start_price = None  # Track start of current mid-rise
        self.in_mid_rise = False  # Track if currently in a mid-rise (consecutive up days)
        
        # Event cumulative percentage tracking (for CSV export)
        self.event_cumulative_pct = 0  # Net cumulative percentage during current event
        self.event_last_price = None  # Track last price for cumulative calculation
        
    def update_phase(self, current_price: float, prev_price: float, date: datetime):
        """Update market phase using FIRST DIP → RECOVERY → SECOND DIP pattern for fall detection.
        
        RISE detection: 2 consecutive up days
        FALL detection: First dip → recovery → second dip (trace back to peak for fall start)
        """
        # Track price history (keep last 4 days for rise start detection with hot stock)
        self.price_history.append((date, current_price))
        if len(self.price_history) > 4:
            self.price_history.pop(0)
        
        # Treat ±$0.01 as plateau (negligible movement)
        price_diff = abs(current_price - prev_price)
        is_plateau = price_diff <= 0.01
        is_up = current_price > prev_price and not is_plateau
        is_down = current_price < prev_price and not is_plateau
        
        # Track consecutive up days for rise detection
        if is_up:
            self.consecutive_up_days += 1
        else:
            self.consecutive_up_days = 0
        
        if date.year >= 2025 and ((date.month == 5 and date.day >= 20) or date.month == 6):
            print(f"  🐛 {date.strftime('%Y-%m-%d')}: phase={self.phase}, is_up={is_up}, consecutive_up={self.consecutive_up_days}, ${prev_price:.2f}→${current_price:.2f}")
        
        # Debug for Sept 2022 to see why we get stuck
        if date.year == 2022 and date.month == 9 and date.day >= 12:
            print(f"  🐛 {date.strftime('%Y-%m-%d')}: phase={self.phase}, is_up={is_up}, is_down={is_down}, consecutive_up={self.consecutive_up_days}, ${prev_price:.2f}→${current_price:.2f}")
        
        if self.phase == MarketPhase.UNKNOWN or self.phase == MarketPhase.FALLING:
            # HUNTING FOR RISE START: Need 2 consecutive up days (or 3 for hot stocks)
            # Check if we're in a hot stock situation (shopping spree)
            is_hot_stock = bool(self.insiders_bought_in_rise and self.insiders_bought_in_fall)
            required_up_days = 3 if is_hot_stock else 2
            
            if self.consecutive_up_days >= required_up_days:
                # Temporarily disable 15-day gap to test
                days_since_last_peak = 999
                
                if date.year >= 2025:
                    print(f"  🔍 {required_up_days} consecutive up days on {date.strftime('%Y-%m-%d')}, days_since_last_peak: {days_since_last_peak}, hot_stock: {is_hot_stock}")
                
                if days_since_last_peak >= 0:  # Always allow (testing)
                    # Clear old insider data only if there hasn't been recent activity (30+ days)
                    if not self.in_position:
                        days_since_last_purchase = 999
                        for purchase_list in [self.insiders_bought_in_rise, self.insiders_bought_in_fall]:
                            for purchase in purchase_list:
                                purchase_date = pd.to_datetime(purchase['date']).tz_localize(None)
                                current_date_naive = date.tz_localize(None) if hasattr(date, 'tz_localize') else date
                                days_diff = (current_date_naive - purchase_date).days
                                days_since_last_purchase = min(days_since_last_purchase, days_diff)
                        
                        if days_since_last_purchase > 30:
                            self.insiders_bought_in_rise = []
                            self.insiders_bought_in_fall = []
                            self.shopping_spree_peak_price = None
                    
                    # Find the actual rise start (the bottom before the climb)
                    # For 2-day detection: bottom is 3 days ago
                    # For 3-day detection (hot stock): bottom is 4 days ago
                    lookback = 4 if is_hot_stock else 3
                    
                    if len(self.price_history) >= lookback:
                        bottom_date, bottom_price = self.price_history[-lookback]
                        actual_start_date = bottom_date
                        actual_start_price = bottom_price
                    else:
                        # Fallback if not enough history
                        actual_start_date = date
                        actual_start_price = prev_price
                    
                    # Record the completed FALL event if we were falling before
                    if self.phase == MarketPhase.FALLING and self.trend_start_date:
                        fall_days = (actual_start_date - self.trend_start_date).days
                        # Use cumulative percentage instead of simple price change
                        fall_pct = self.event_cumulative_pct
                        
                        # SAVE THE FALL PERCENTAGE for buy signal evaluation
                        # But DON'T overwrite if we're already in a position (target is locked in)
                        if not self.in_position:
                            self.prev_fall_pct = abs(fall_pct)
                            self.prev_fall_start_price = self.trend_start_price
                        
                        # Filter insiders who bought during THIS fall only
                        # Make dates tz-naive for comparison
                        fall_start_naive = self.trend_start_date.tz_localize(None) if hasattr(self.trend_start_date, 'tz_localize') else self.trend_start_date
                        fall_end_naive = actual_start_date.tz_localize(None) if hasattr(actual_start_date, 'tz_localize') else actual_start_date
                        
                        fall_insiders = [
                            i for i in self.insiders_bought_in_fall 
                            if fall_start_naive <= pd.to_datetime(i['date']).tz_localize(None) <= fall_end_naive
                        ]
                        
                        self.all_events.append({
                            'event_type': 'DOWN',
                            'start_date': self.trend_start_date,
                            'end_date': actual_start_date,
                            'days': fall_days,
                            'change_pct': fall_pct,
                            'start_price': self.trend_start_price,
                            'end_price': self.trend_low_price,
                            'insiders': fall_insiders
                        })
                    
                    # Start rise: Reset peak tracking for this NEW rise cycle
                    self.phase = MarketPhase.RISING
                    self.trend_start_date = actual_start_date  # Rise started at the bottom
                    self.trend_start_price = actual_start_price
                    self.trend_peak_price = current_price  # Peak for THIS rise starts here
                    self.trend_peak_date = date
                    self.first_dip_date = None  # Reset fall detection
                    self.first_dip_price = None
                    self.in_recovery = False
                    # Reset cumulative percentage tracking for new rise event
                    self.event_cumulative_pct = 0
                    self.event_last_price = actual_start_price
                    print(f"📈 RISE STARTED on {actual_start_date.strftime('%Y-%m-%d')} at ${actual_start_price:.2f} (detected on {date.strftime('%Y-%m-%d')}): peak tracking reset to ${current_price:.2f}")
        
        elif self.phase == MarketPhase.RISING:
            # Track cumulative percentage change during rise
            if self.event_last_price and self.event_last_price > 0:
                daily_pct = ((current_price - self.event_last_price) / self.event_last_price) * 100
                self.event_cumulative_pct += daily_pct
            self.event_last_price = current_price
            
            # Update peak if new high
            if current_price > self.trend_peak_price:
                old_peak = self.trend_peak_price
                self.trend_peak_price = current_price
                self.trend_peak_date = date
                # Reset fall detection when making new highs
                self.first_dip_date = None
                self.first_dip_price = None
                self.dip_low_price = None
                self.in_recovery = False
                if date.year >= 2025 and date.month >= 5:
                    print(f"  🔍 New peak: ${old_peak:.2f} → ${current_price:.2f} on {date.strftime('%Y-%m-%d')}")
            
            # FALL DETECTION: First dip → recovery → second dip
            if is_down:
                decline_from_peak = ((self.trend_peak_price - current_price) / self.trend_peak_price) * 100
                if date.year >= 2025 and date.month >= 5:
                    print(f"  🔍 Down day: peak=${self.trend_peak_price:.2f}, current=${current_price:.2f}, decline={decline_from_peak:.1f}%, first_dip={self.first_dip_date}, in_recovery={self.in_recovery}")
                
                if self.first_dip_date is None:
                    # This is the FIRST DIP - remember it
                    # Also remember the day BEFORE the dip (yesterday) as the actual end of the rise
                    if decline_from_peak >= 1.0:  # Meaningful dip (≥1%)
                        self.first_dip_date = date
                        self.first_dip_price = current_price
                        self.dip_low_price = current_price  # Start tracking lowest point
                        self.in_recovery = False
                        # The rise actually ended yesterday (the last day before this dip)
                        # We'll update trend_peak_date to yesterday when we confirm the fall
                        if date.year >= 2025:
                            print(f"  🔍 First dip detected on {date.strftime('%Y-%m-%d')}: ${self.trend_peak_price:.2f} → ${current_price:.2f} (-{decline_from_peak:.1f}%)")
                elif not self.in_recovery:
                    # Still dipping - update lowest point
                    if current_price < self.dip_low_price:
                        self.dip_low_price = current_price
                        if date.year >= 2025:
                            print(f"  🔍 New dip low on {date.strftime('%Y-%m-%d')}: ${current_price:.2f}")
                elif self.in_recovery:
                    # We had first dip, then recovery, now checking for SECOND DIP
                    # Require meaningful dip (>1% from recovery high), not just $0.01 movement
                    decline_from_recovery = ((self.recovery_high - current_price) / self.recovery_high) * 100
                    if decline_from_recovery >= 1.0:
                        # Meaningful second dip - FALL CONFIRMED
                        if date.year >= 2025:
                            print(f"  🔍 Second dip detected on {date.strftime('%Y-%m-%d')}: ${self.recovery_high:.2f} → ${current_price:.2f} (-{decline_from_recovery:.1f}%) - FALL CONFIRMED")
                        self.prev_rise_pct = ((self.trend_peak_price - self.trend_start_price) / 
                                             self.trend_start_price) * 100
                        self.prev_rise_start_price = self.trend_start_price
                        self.prev_rise_peak_price = self.trend_peak_price
                        self.last_peak_date = self.trend_peak_date
                        
                        # Transition to FALLING
                        self.phase = MarketPhase.FALLING
                        
                        # The rise actually ended the day BEFORE the first dip
                        # Use business day offset to get previous trading day
                        from pandas.tseries.offsets import BDay
                        actual_rise_end = self.first_dip_date - BDay(1)
                        
                        # Record the completed RISE event before transitioning
                        rise_days = (actual_rise_end - self.trend_start_date).days
                        # Use cumulative percentage instead of simple price change
                        rise_pct = self.event_cumulative_pct
                        
                        # Filter insiders who bought during THIS rise only
                        # Make dates tz-naive for comparison
                        rise_start_naive = self.trend_start_date.tz_localize(None) if hasattr(self.trend_start_date, 'tz_localize') else self.trend_start_date
                        rise_end_naive = actual_rise_end.tz_localize(None) if hasattr(actual_rise_end, 'tz_localize') else actual_rise_end
                        
                        rise_insiders = [
                            i for i in self.insiders_bought_in_rise 
                            if rise_start_naive <= pd.to_datetime(i['date']).tz_localize(None) <= rise_end_naive
                        ]
                        
                        self.all_events.append({
                            'event_type': 'RISE',
                            'start_date': self.trend_start_date,
                            'end_date': actual_rise_end,
                            'days': rise_days,
                            'change_pct': rise_pct,
                            'start_price': self.trend_start_price,
                            'end_price': None,
                            'peak_price': self.trend_peak_price,
                            'insiders': rise_insiders
                        })
                        
                        self.trend_start_date = self.first_dip_date  # Fall starts from first dip day
                        self.trend_start_price = self.trend_peak_price
                        self.trend_low_price = current_price
                        self.trend_low_date = date
                        self.consecutive_up_days = 0
                        # Reset cumulative percentage tracking for new fall event
                        self.event_cumulative_pct = 0
                        self.event_last_price = self.trend_peak_price
                        print(f"📉 FALL STARTED on {self.trend_peak_date.strftime('%Y-%m-%d')} (detected on {date.strftime('%Y-%m-%d')} via dip→recovery→dip)")
                        
                        # Reset fall detection
                        self.first_dip_date = None
                        self.first_dip_price = None
                        self.dip_low_price = None
                        self.in_recovery = False
                    else:
                        # Not meaningful enough - ignore this tiny dip
                        if date.year >= 2025:
                            print(f"  🔍 Tiny dip ignored on {date.strftime('%Y-%m-%d')}: only -{decline_from_recovery:.2f}%")
            
            # If stock goes UP after recovery, reset fall detection (rise continues!)
            elif is_up and self.in_recovery:
                if date.year >= 2025:
                    print(f"  🔍 Stock went up after recovery on {date.strftime('%Y-%m-%d')}: ${prev_price:.2f} → ${current_price:.2f} - resetting fall detection, rise continues!")
                # Reset fall detection
                self.first_dip_date = None
                self.first_dip_price = None
                self.dip_low_price = None
                self.in_recovery = False
                # Update recovery high
                self.recovery_high = current_price
            
            # RECOVERY DETECTION (up day OR plateau after first dip)
            elif (is_up or is_plateau) and self.first_dip_date is not None and not self.in_recovery:
                # Price going up OR plateau (same price) after first dip - this is RECOVERY!
                recovery_from_dip = ((current_price - self.dip_low_price) / self.dip_low_price) * 100
                if recovery_from_dip >= 0.0:  # Any non-decline is recovery (plateau counts!)
                    self.in_recovery = True
                    self.recovery_high = current_price
                    if date.year >= 2025:
                        print(f"  🔍 Recovery detected on {date.strftime('%Y-%m-%d')}: ${self.dip_low_price:.2f} → ${current_price:.2f} (+{recovery_from_dip:.1f}%)")
        
        # Track lowest price during fall
        if self.phase == MarketPhase.FALLING:
            # Track cumulative percentage change during fall
            if self.event_last_price and self.event_last_price > 0:
                daily_pct = ((current_price - self.event_last_price) / self.event_last_price) * 100
                self.event_cumulative_pct += daily_pct
            self.event_last_price = current_price
            
            if current_price < self.trend_low_price:
                self.trend_low_price = current_price
                self.trend_low_date = date
        
        # Calculate current fall percentage
        # DON'T update prev_fall_pct if we're in a position (target is locked in)
        if self.phase == MarketPhase.FALLING and self.trend_start_price and not self.in_position:
            self.prev_fall_pct = ((self.trend_start_price - current_price) / self.trend_start_price) * 100
    
    def record_insider_purchase(self, date_str: str, trade_info: Dict):
        """Record an insider purchase occurring today."""
        trade_data = {
            'date': date_str,
            'price': trade_info['price'],
            'insider_name': trade_info['insider_name'],
            'value': trade_info['value'],
            'stock_price': trade_info.get('stock_price', trade_info['price'])  # Track stock price on this day
        }
        
        if self.phase == MarketPhase.RISING:
            self.insiders_bought_in_rise.append(trade_data)
            # Track peak stock price during shopping spree
            if self.shopping_spree_peak_price is None or trade_data['stock_price'] > self.shopping_spree_peak_price:
                old_peak = self.shopping_spree_peak_price
                self.shopping_spree_peak_price = trade_data['stock_price']
                print(f"    📈 Shopping spree peak updated: ${old_peak if old_peak else 0:.2f} → ${self.shopping_spree_peak_price:.2f}")
        elif self.phase == MarketPhase.FALLING:
            self.insiders_bought_in_fall.append(trade_data)
            # Continue tracking peak during fall phase of shopping spree
            if self.shopping_spree_peak_price is None or trade_data['stock_price'] > self.shopping_spree_peak_price:
                old_peak = self.shopping_spree_peak_price
                self.shopping_spree_peak_price = trade_data['stock_price']
                print(f"    📈 Shopping spree peak updated: ${old_peak if old_peak else 0:.2f} → ${self.shopping_spree_peak_price:.2f}")
    
    def check_buy_signal(self, current_date: datetime, current_price: float) -> Optional[Dict]:
        """Check if we should buy based on current state (NO HINDSIGHT).
        
        For shopping spree (hot stocks), require 3 consecutive up days for confirmation.
        """
        
        if self.in_position:
            return None
        
        # We need to be in a rising phase (recovering from fall)
        if self.phase != MarketPhase.RISING:
            return None
        
        # No insider activity? No trade
        if not self.insiders_bought_in_fall:
            return None
        
        # For HOT STOCKS (shopping spree), require 3 consecutive up days
        # 2 days detect the rise, but we wait for the 3rd day to confirm before buying
        if self.insiders_bought_in_rise and self.insiders_bought_in_fall:
            # This is a shopping spree - be conservative
            if self.consecutive_up_days < 3:
                return None  # Wait for 3rd consecutive up day
        
        # SCENARIO 1: Shopping Spree
        # - Insiders bought during the rise
        # - Continued buying during the fall  
        # - Target = highest STOCK price during the shopping spree period
        if self.insiders_bought_in_rise and self.insiders_bought_in_fall and self.shopping_spree_peak_price:
            # Use the peak stock price during shopping spree as target
            target = self.shopping_spree_peak_price
            
            if target > current_price:
                self.in_position = True
                self.entry_date = current_date
                self.entry_price = current_price
                self.target_price = target
                self.buy_type = 'shopping_spree'
                self.target_reached = False
                self.max_mid_fall_before_target = 0
                self.peak_since_entry = 0
                
                all_insider_purchases = self.insiders_bought_in_rise + self.insiders_bought_in_fall
                return {
                    'buy_date': current_date,
                    'entry_price': current_price,
                    'target_price': target,
                    'buy_type': 'shopping_spree',
                    'num_insiders': len(all_insider_purchases),
                    'prev_fall_pct': self.prev_fall_pct
                }
        
        # SCENARIO 2: Absorption Buy (insiders buy ONLY during fall, minimum $5K)
        elif self.insiders_bought_in_fall and not self.insiders_bought_in_rise:
            # Calculate total insider investment during the fall
            total_investment = sum(abs(t['value']) for t in self.insiders_bought_in_fall)
            
            # Require at least $5,000 invested
            if total_investment >= 5000:
                # Target is to recover what was lost (fall magnitude)
                target_gain_pct = abs(self.prev_fall_pct)
                
                self.in_position = True
                self.entry_date = current_date
                self.entry_price = current_price
                self.target_price = current_price * (1 + target_gain_pct / 100)
                self.buy_type = 'absorption_buy'
                self.target_reached = False
                self.peak_since_entry = 0
                # Use trend_start_price (actual rise start) for cumulative calculation
                self.rise_start_price = self.trend_start_price if self.trend_start_price else current_price
                self.in_mid_rise = False  # Track if currently in a mid-rise
                
                return {
                    'buy_date': current_date,
                    'entry_price': current_price,
                    'target_price': self.target_price,
                    'buy_type': 'absorption_buy',
                    'num_insiders': len(self.insiders_bought_in_fall),
                    'fall_pct': self.prev_fall_pct,
                    'total_investment': total_investment,
                    'rise_start_price': self.rise_start_price  # Debug
                }
        
        # Clear insider data after checking - we've evaluated this rise+fall cycle
        # This prevents stale data from affecting future signals
        # ALWAYS clear after checking (whether we bought or not)
        if self.insiders_bought_in_rise or self.insiders_bought_in_fall:
            print(f"    🧹 Clearing insider data (rise: {len(self.insiders_bought_in_rise)}, fall: {len(self.insiders_bought_in_fall)}, peak: ${self.shopping_spree_peak_price or 0:.2f})")
        self.insiders_bought_in_rise = []
        self.insiders_bought_in_fall = []
        self.shopping_spree_peak_price = None
        
        return None
    
    def check_sell_signal(self, current_date: datetime, current_price: float, 
                         prev_price: float) -> Optional[Tuple[str, float]]:
        """Check if we should sell based on current conditions (NO HINDSIGHT)."""
        
        if not self.in_position:
            return None
        
        # Calculate current performance
        current_gain_pct = ((current_price - self.entry_price) / self.entry_price) * 100
        self.peak_since_entry = max(self.peak_since_entry, current_gain_pct)
        
        # ABSORPTION BUY has different sell logic
        if self.buy_type == 'absorption_buy':
            if self.phase == MarketPhase.RISING:
                # Track MID-RISES within the current rise event
                # Accumulate all mid-rise percentages and check if cumulative >= fall_pct
                daily_change_pct = ((current_price - prev_price) / prev_price) * 100
                
                # Start a new mid-rise when going up
                if daily_change_pct > 0:
                    if not self.in_mid_rise:
                        # Starting a new mid-rise
                        self.in_mid_rise = True
                        self.mid_rise_start_price = prev_price
                # End mid-rise when going down
                elif daily_change_pct < -1.0 and self.in_mid_rise:
                    # Mid-rise ended - calculate and add to cumulative
                    mid_rise_pct = ((prev_price - self.mid_rise_start_price) / self.mid_rise_start_price) * 100
                    if mid_rise_pct >= 1.0:  # Only count meaningful mid-rises
                        self.cumulative_mid_rises_pct += mid_rise_pct
                        print(f"  📈 Mid-rise ended: +{mid_rise_pct:.2f}%, cumulative: {self.cumulative_mid_rises_pct:.2f}%")
                    self.in_mid_rise = False
                    self.mid_rise_start_price = None
                
                # Check if cumulative mid-rises reached target
                target_gain_pct = abs(self.prev_fall_pct)
                if not self.target_reached and self.cumulative_mid_rises_pct >= target_gain_pct:
                    self.target_reached = True
                    print(f"  🎯 Target reached! Cumulative mid-rises: {self.cumulative_mid_rises_pct:.2f}% >= {target_gain_pct:.2f}%")
                
                # After target reached, sell on first dip
                if self.target_reached and daily_change_pct < -1.0:
                    self.in_position = False
                    return ('absorption_target_reached', current_price)
            
            # No stop loss for absorption buy
            return None
        
        # SHOPPING SPREE sell logic (original)
        # Emergency stop loss: -15%
        if current_gain_pct <= -15:
            self.in_position = False
            return ('stop_loss', current_price)
        
        # Check if target reached
        if not self.target_reached and current_price >= self.target_price:
            self.target_reached = True
        
        # Track mid-falls before reaching target
        daily_change_pct = ((current_price - prev_price) / prev_price) * 100
        if not self.target_reached and daily_change_pct < -1.0:
            self.max_mid_fall_before_target = max(self.max_mid_fall_before_target, 
                                                   abs(daily_change_pct))
        
        # SELL CONDITIONS after target reached
        if self.target_reached:
            # For ALL scenarios: Sell on first dip after reaching target
            # Don't sell during rises - wait for the dip
            if daily_change_pct < -1.0:
                self.in_position = False
                return ('target_reached_first_dip', current_price)
        
        return None


def simulate_live_trading(insider_trades: Dict, price_df: pd.DataFrame, ticker: str = 'UNKNOWN') -> List[Dict]:
    """
    Simulate live trading day-by-day with NO HINDSIGHT.
    
    Process each day sequentially, making decisions based only on past data.
    """
    state = TradingState()
    completed_trades = []
    
    print("\n🔴 LIVE TRADING SIMULATION (No Hindsight)")
    print("=" * 80)
    
    # Process day by day
    for i in range(1, len(price_df)):
        current_date = price_df.index[i]
        current_price = price_df['Close'].iloc[i]
        prev_price = price_df['Close'].iloc[i-1]
        date_str = current_date.strftime('%Y-%m-%d')
        
        # Update market phase
        state.update_phase(current_price, prev_price, current_date)
        
        # Check for insider purchases today
        if date_str in insider_trades:
            # Handle multiple insiders trading on the same day
            for trade_info in insider_trades[date_str]:
                # Add current stock price to the trade info
                trade_info_with_price = trade_info.copy()
                trade_info_with_price['stock_price'] = current_price
                state.record_insider_purchase(date_str, trade_info_with_price)
                insider_name = trade_info['insider_name']
                print(f"  📢 {date_str}: Insider purchase detected - {insider_name} @ ${current_price:.2f}")
        
        # If we're in a position, check for sell signal
        if state.in_position:
            sell_signal = state.check_sell_signal(current_date, current_price, prev_price)
            if sell_signal:
                reason, exit_price = sell_signal
                days_held = (current_date - state.entry_date).days
                return_pct = ((exit_price - state.entry_price) / state.entry_price) * 100
                profit = state.position_size * (return_pct / 100)
                
                trade = {
                    'ticker': ticker,
                    'entry_date': state.entry_date.strftime('%Y-%m-%d'),
                    'entry_price': round(state.entry_price, 2),
                    'exit_date': current_date.strftime('%Y-%m-%d'),
                    'exit_price': round(exit_price, 2),
                    'target_price': round(state.target_price, 2),
                    'days_held': days_held,
                    'return_pct': round(return_pct, 2),
                    'position_size': state.position_size,
                    'profit_loss': round(profit, 2),
                    'sell_reason': reason,
                    'target_reached': 'yes' if state.target_reached else 'no',
                    'peak_gain': round(state.peak_since_entry, 2),
                    'buy_type': state.buy_type
                }
                
                completed_trades.append(trade)
                status = "✓ TARGET" if state.target_reached else "✗ NO TARGET"
                print(f"  💰 SELL {date_str}: {return_pct:+.1f}% ({days_held}d) {status} - {reason}")
                
                # Reset insider tracking AND peak price after trade completes
                state.insiders_bought_in_fall = []
                state.insiders_bought_in_rise = []
                state.shopping_spree_peak_price = None
                print(f"🧹 Trade completed - clearing all insider data and peak price")
        
        # If not in position, check for buy signal
        else:
            buy_signal = state.check_buy_signal(current_date, current_price)
            if buy_signal:
                print(f"  🟢 BUY {date_str}: {buy_signal['buy_type'].upper()} @ ${current_price:.2f} "
                      f"(Target: ${buy_signal['target_price']:.2f}, Insiders: {buy_signal['num_insiders']}, "
                      f"Fall: {buy_signal.get('fall_pct', 0):.2f}%, Rise Start: ${buy_signal.get('rise_start_price', 0):.2f})")
    
    # Handle open position at end
    if state.in_position:
        final_date = price_df.index[-1]
        final_price = price_df['Close'].iloc[-1]
        days_held = (final_date - state.entry_date).days
        return_pct = ((final_price - state.entry_price) / state.entry_price) * 100
        profit = state.position_size * (return_pct / 100)
        
        trade = {
            'ticker': ticker,
            'entry_date': state.entry_date.strftime('%Y-%m-%d'),
            'entry_price': round(state.entry_price, 2),
            'exit_date': final_date.strftime('%Y-%m-%d'),
            'exit_price': round(final_price, 2),
            'target_price': round(state.target_price, 2),
            'days_held': days_held,
            'return_pct': round(return_pct, 2),
            'position_size': state.position_size,
            'profit_loss': round(profit, 2),
            'sell_reason': 'end_of_period',
            'target_reached': 'yes' if state.target_reached else 'no',
            'peak_gain': round(state.peak_since_entry, 2),
            'buy_type': state.buy_type
        }
        
        completed_trades.append(trade)
        print(f"  💰 SELL {final_date.strftime('%Y-%m-%d')}: {return_pct:+.1f}% - end_of_period")
    
    print("=" * 80)
    return completed_trades, state  # Return state to access all_events


def generate_event_files(events: List[Dict], price_df: pd.DataFrame, ticker: str = 'BSFC'):
    """Generate CSV and Excel files for rise/fall events with proper formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from datetime import datetime
    
    if not events:
        print("⚠️ No events to export")
        return
    
    # Look up actual end prices for RISE events (FALL events already have end_price)
    for event in events:
        if event['end_price'] is None:
            try:
                end_date = event['end_date']
                end_price = price_df.loc[end_date]['Close']
                event['end_price'] = end_price
            except:
                # Fallback: use peak price if date not found
                event['end_price'] = event.get('peak_price', event['start_price'])
    
    # Calculate cumulative percentages and ranks
    cumulative_pct = 0
    rise_events = [e for e in events if e['event_type'] == 'RISE']
    fall_events = [e for e in events if e['event_type'] == 'DOWN']
    
    # Rank rises and falls by change_pct
    rise_events_sorted = sorted(rise_events, key=lambda x: abs(x['change_pct']), reverse=True)
    fall_events_sorted = sorted(fall_events, key=lambda x: abs(x['change_pct']), reverse=True)
    
    rise_ranks = {id(e): f"{i+1}/{len(rise_events)}" for i, e in enumerate(rise_events_sorted)}
    fall_ranks = {id(e): f"{i+1}/{len(fall_events)}" for i, e in enumerate(fall_events_sorted)}
    
    # Prepare data for export
    export_data = []
    for event in events:
        cumulative_pct += event['change_pct']
        
        # Format insider purchases
        insiders_str = ""
        if event['insiders']:
            insider_dates = sorted(set([i['date'] for i in event['insiders']]))
            insiders_str = ", ".join([datetime.strptime(d, '%Y-%m-%d').strftime('%d/%m/%Y') for d in insider_dates])
        
        # Get rank
        rank = rise_ranks.get(id(event)) if event['event_type'] == 'RISE' else fall_ranks.get(id(event))
        
        export_data.append({
            'event_type': event['event_type'],
            'start_date': event['start_date'].strftime('%d/%m/%Y'),
            'end_date': event['end_date'].strftime('%d/%m/%Y'),
            'days': event['days'],
            'change_pct': round(event['change_pct'], 2),
            'cumulative_pct': round(cumulative_pct, 2),
            'insider_purchases': insiders_str if insiders_str else None,
            'rank': rank
        })
    
    # Save to CSV
    csv_file = f'output CSVs/{ticker.lower()}_rise_events.csv'
    df = pd.DataFrame(export_data)
    df.to_csv(csv_file, index=False)
    print(f"✓ CSV saved to: {csv_file}")
    
    # Create Excel with colors
    wb = Workbook()
    ws = wb.active
    ws.title = f"{ticker} Rise-Fall Events"
    
    # Colors
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray
    rise_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")    # Light green
    fall_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")    # Light pink
    
    # Headers
    headers = ['Event Type', 'Start Date', 'End Date', 'Days', 'Change %', 'Rank', 'Cumulative %', 'Insider Purchases']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
    
    # Data rows
    for row_data in export_data:
        ws.append([
            row_data['event_type'],
            row_data['start_date'],
            row_data['end_date'],
            row_data['days'],
            row_data['change_pct'],
            row_data['rank'],
            row_data['cumulative_pct'],
            row_data['insider_purchases']
        ])
        
        # Apply color based on event type
        fill = rise_fill if row_data['event_type'] == 'RISE' else fall_fill
        for cell in ws[ws.max_row]:
            cell.fill = fill
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 30
    
    # Save Excel
    excel_file = f'output CSVs/{ticker.lower()}_rise_events.xlsx'
    wb.save(excel_file)
    print(f"✓ Excel saved to: {excel_file}")


def analyze_rise_volatility(df: pd.DataFrame, rise_event: Dict) -> Dict:
    """
    Analyze the volatility pattern during and after a rise event.
    Tracks mid-rises, mid-falls during the rise, and post-rise behavior.
    
    Args:
        df: DataFrame with Date index and Close prices
        rise_event: Dictionary containing start_date, end_date, and change_pct
        
    Returns:
        Dictionary with detailed volatility metrics
    """
    # Get dates directly - they're already Timestamps
    start_date = rise_event['start_date']
    end_date = rise_event['end_date']
    
    # Get price data for the rise period
    rise_df = df.loc[start_date:end_date]
    
    # Get post-rise data - track for up to 30 trading days
    # Use iloc for robust indexing since some dates might not exist in DataFrame
    try:
        end_idx = df.index.get_indexer([end_date], method='nearest')[0]
    except:
        # If end_date not found, skip this event
        return None
        
    post_rise_end_idx = min(end_idx + 31, len(df))
    post_rise_df = df.iloc[end_idx:post_rise_end_idx]
    
    result = {
        'rise_start_date': start_date.strftime('%d/%m/%Y'),
        'rise_end_date': end_date.strftime('%d/%m/%Y'),
        'rise_days': rise_event['days'],
        'rise_percentage': rise_event['change_pct'],
        'mid_rises': {},
        'mid_falls': {},
        'first_dip': None,
        'first_recovery': None,
        'second_dip': None,
        'insiders': rise_event.get('insiders', [])
    }
    
    # Track consecutive movements during the rise - group consecutive rises/falls
    if len(rise_df) > 1:
        i = 0
        while i < len(rise_df) - 1:
            # Start tracking a potential movement
            movement_start_price = rise_df['Close'].iloc[i]
            movement_start_idx = i
            current_direction = None  # 'up' or 'down'
            
            # Keep going in the same direction until it changes
            j = i + 1
            while j < len(rise_df):
                prev_price = rise_df['Close'].iloc[j-1]
                current_price = rise_df['Close'].iloc[j]
                
                # Determine direction of this day
                if current_price > prev_price:
                    day_direction = 'up'
                elif current_price < prev_price:
                    day_direction = 'down'
                else:
                    # No change, continue in same direction
                    j += 1
                    continue
                
                # First day sets the direction
                if current_direction is None:
                    current_direction = day_direction
                    j += 1
                    continue
                
                # If direction changed, we've completed a movement
                if day_direction != current_direction:
                    break
                    
                j += 1
            
            # Calculate the total movement from start to end
            if current_direction is not None and j > i + 1:
                movement_end_price = rise_df['Close'].iloc[j-1]
                movement_end_date = rise_df.index[j-1]
                total_change_pct = ((movement_end_price - movement_start_price) / movement_start_price) * 100
                
                # Record rises ≥1%
                if current_direction == 'up' and total_change_pct >= 1.0:
                    pct_key = str(round(total_change_pct, 2))
                    result['mid_rises'][pct_key] = {
                        'date': movement_end_date.strftime('%d/%m/%Y')
                    }
                # Record falls ≥1%
                elif current_direction == 'down' and total_change_pct <= -1.0:
                    pct_key = str(round(total_change_pct, 2))
                    result['mid_falls'][pct_key] = {
                        'date': movement_end_date.strftime('%d/%m/%Y')
                    }
                
                # Move to the end of this movement
                i = j - 1
            else:
                i += 1
    
    # NEW LOGIC: Check for 2 consecutive declining mid-rises (indicates this should be a DOWN event)
    # If we detect: mid_rise at price X, mid_rise at price Y (Y < X), mid_rise at price Z (Z < Y)
    # Then this is actually a DOWN event starting from the first mid_rise
    if len(result['mid_rises']) >= 2:
        # Get mid-rises sorted by date
        mid_rises_with_prices = []
        for pct_key, rise_info in result['mid_rises'].items():
            date_str = rise_info['date']
            rise_date = pd.to_datetime(date_str, format='%d/%m/%Y')
            # Find the price at this date
            try:
                price_at_date = rise_df.loc[rise_date, 'Close']
                mid_rises_with_prices.append({
                    'date': rise_date,
                    'date_str': date_str,
                    'price': price_at_date,
                    'pct': pct_key
                })
            except:
                pass
        
        # Sort by date
        mid_rises_with_prices.sort(key=lambda x: x['date'])
        
        # Check for 2 consecutive declines
        detected_fall_start = None
        for i in range(len(mid_rises_with_prices) - 2):
            price1 = mid_rises_with_prices[i]['price']
            price2 = mid_rises_with_prices[i + 1]['price']
            price3 = mid_rises_with_prices[i + 2]['price']
            
            # Check if we have 2 consecutive declines: price2 < price1 AND price3 < price2
            if price2 < price1 and price3 < price2:
                # This is a fall event starting from the first mid-rise
                detected_fall_start = mid_rises_with_prices[i]
                print(f"  ⚠️  DETECTED FALL PATTERN: {price1:.2f} -> {price2:.2f} -> {price3:.2f}")
                print(f"  ⚠️  Reclassifying as DOWN event starting from {detected_fall_start['date_str']}")
                
                # Reclassify this as a DOWN event
                # The rise actually ended at the first mid-rise where the decline started
                result['rise_end_date'] = detected_fall_start['date_str']
                
                # Recalculate rise percentage using the corrected end date
                try:
                    corrected_end_date = detected_fall_start['date']
                    rise_start_price = rise_df['Close'].iloc[0]
                    corrected_end_price = rise_df.loc[corrected_end_date, 'Close']
                    corrected_rise_pct = ((corrected_end_price - rise_start_price) / rise_start_price) * 100
                    result['rise_percentage'] = corrected_rise_pct
                    
                    # Recalculate days
                    corrected_rise_days = (corrected_end_date - start_date).days
                    result['rise_days'] = corrected_rise_days
                    
                    print(f"  ✓ Corrected: {result['rise_start_date']} to {result['rise_end_date']} = {corrected_rise_pct:.2f}% over {corrected_rise_days} days")
                except Exception as e:
                    print(f"  ✗ Error correcting rise event: {e}")
                
                break
    
    # Analyze post-rise behavior: Track first dip -> first recovery -> second dip pattern
    if len(post_rise_df) > 1:
        peak_price = post_rise_df['Close'].iloc[0]
        
        # State tracking
        first_dip_found = False
        first_dip_low = peak_price
        first_dip_low_idx = 0
        
        first_recovery_found = False
        recovery_high = peak_price
        recovery_high_idx = 0
        
        for i in range(1, len(post_rise_df)):
            current_price = post_rise_df['Close'].iloc[i]
            current_date = post_rise_df.index[i]
            
            if not first_dip_found:
                # Looking for the first dip
                if current_price < first_dip_low:
                    first_dip_low = current_price
                    first_dip_low_idx = i
                    fall_pct = ((peak_price - current_price) / peak_price) * 100
                    if fall_pct >= 1.0:
                        result['first_dip'] = {
                            'date': current_date.strftime('%d/%m/%Y'),
                            'percentage': round(-fall_pct, 2),
                            'days_after_peak': i
                        }
                # Check if price is recovering
                elif current_price > first_dip_low and result['first_dip'] is not None:
                    recovery_from_low = ((current_price - first_dip_low) / first_dip_low) * 100
                    if recovery_from_low >= 1.0:
                        first_dip_found = True
                        recovery_high = current_price
                        recovery_high_idx = i
                        result['first_recovery'] = {
                            'date': current_date.strftime('%d/%m/%Y'),
                            'percentage': round(recovery_from_low, 2),
                            'days_after_peak': i
                        }
                    
            elif not first_recovery_found:
                # Looking for the first recovery
                if current_price > recovery_high:
                    recovery_high = current_price
                    recovery_high_idx = i
                    recovery_pct = ((current_price - first_dip_low) / first_dip_low) * 100
                    if recovery_pct >= 1.0:
                        result['first_recovery'] = {
                            'date': current_date.strftime('%d/%m/%Y'),
                            'percentage': round(recovery_pct, 2),
                            'days_after_peak': i
                        }
                # Check if price is declining again
                elif current_price < recovery_high and result['first_recovery'] is not None:
                    first_recovery_found = True
                    
            else:
                # Looking for the second dip
                second_dip_pct = ((recovery_high - current_price) / recovery_high) * 100
                if second_dip_pct >= 1.0:
                    days_since_recovery = i - recovery_high_idx
                    if result['second_dip'] is None or second_dip_pct > abs(result['second_dip']['percentage']):
                        result['second_dip'] = {
                            'date': current_date.strftime('%d/%m/%Y'),
                            'percentage': round(-second_dip_pct, 2),
                            'days_after_peak': i,
                            'days_since_recovery': days_since_recovery
                        }
    
    return result


def generate_volatility_json(events: List[Dict], df: pd.DataFrame, ticker: str = 'BSFC'):
    """Generate JSON with volatility analysis for all rise events."""
    from datetime import datetime
    
    # Filter only RISE events
    rise_events = [e for e in events if e['event_type'] == 'RISE']
    
    # Build volatility analysis for each rise
    volatility_analysis = {}
    for rise_event in rise_events:
        analysis = analyze_rise_volatility(df, rise_event)
        
        # Skip if analysis failed (date not found in DataFrame)
        if analysis is None:
            continue
        
        # Use rise percentage as key
        pct_key = str(round(rise_event['change_pct'], 2))
        volatility_analysis[pct_key] = analysis
    
    # Build final JSON structure
    json_output = {
        'ticker': ticker,
        'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'total_rise_events': len(rise_events),
        'rise_events': volatility_analysis
    }
    
    # Save to file
    json_file = f'output CSVs/{ticker.lower()}_rise_volatility_analysis.json'
    with open(json_file, 'w') as f:
        json.dump(json_output, f, indent=2)
    
    print(f"✓ Volatility JSON saved to: {json_file}")


def main():
    """Run the live trading simulation for all stocks."""
    # Load cached yfinance data
    print("📦 Loading cached price data...")
    cache_file = 'output CSVs/yfinance_cache_full.json'
    with open(cache_file, 'r') as f:
        cache = json.load(f)
    
    print(f"   ✅ Loaded cache with {len(cache['data'])} stocks")
    print(f"   📅 Cache created: {cache['metadata']['created']}")
    
    # Convert cache data to pandas DataFrames
    price_cache = {}
    for ticker, ticker_data in cache['data'].items():
        df = pd.DataFrame({
            'Open': ticker_data['open'],
            'High': ticker_data['high'],
            'Low': ticker_data['low'],
            'Close': ticker_data['close'],
            'Volume': ticker_data['volume']
        }, index=pd.to_datetime(ticker_data['dates']))
        price_cache[ticker] = df
    
    # Load all insider trades
    with open('output CSVs/expanded_insider_trades.json', 'r') as f:
        trades_file = json.load(f)
    
    # Extract the data array - it's already grouped by ticker
    stocks_data = trades_file['data']
    
    print()
    print("=" * 80)
    print("INSIDER CONVICTION STRATEGY - LIVE SIMULATION (No Hindsight)")
    print("=" * 80)
    print(f"\n📊 Processing {len(stocks_data)} stocks\n")
    print("=" * 80)
    
    all_results = []
    processed = 0
    
    # Filter to only GROV to regenerate results with correct ticker
    stocks_data = [s for s in stocks_data if s['ticker'] == 'GROV']
    
    for stock_data in stocks_data:
        ticker = stock_data['ticker']
        trades_list = stock_data['trades']
        
        processed += 1
        
        print(f"\n\n{'='*80}")
        print(f"Processing {ticker}...")
        print('='*80)
        
        if not trades_list:
            print(f"⚠️  No trades found for {ticker}, skipping")
            continue
        
        # Try to get cached price data
        try:
            if ticker not in price_cache:
                print(f"⚠️  No cached price data for {ticker}, skipping")
                continue
            
            price_df = price_cache[ticker][['Close']].copy()
            
            if price_df.empty:
                print(f"⚠️  Empty price data for {ticker}, skipping")
                continue
            
            # Convert insider trades to date-indexed dictionary format
            insider_trades = {}  # date -> list of trade info
            for trade in trades_list:
                trade_date = trade['trade_date']
                if trade_date:
                    try:
                        # Parse price (remove $ and commas)
                        price_str = str(trade.get('price', '0')).replace('$', '').replace(',', '')
                        price = float(price_str) if price_str else 0.0
                        
                        # Parse value (remove $, +, commas and convert to float)
                        value_str = str(trade.get('value', '0')).replace('$', '').replace('+', '').replace(',', '')
                        value = float(value_str) if value_str else 0.0
                        
                        trade_info = {
                            'price': price,
                            'insider_name': trade.get('insider_name', ''),
                            'value': value,  # Now numeric
                            'title': trade.get('title', '')
                        }
                        
                        # Store as list to handle multiple insiders on same date
                        if trade_date not in insider_trades:
                            insider_trades[trade_date] = []
                        insider_trades[trade_date].append(trade_info)
                    except:
                        continue
            
            print(f"✓ Loaded {sum(len(v) for v in insider_trades.values())} insider trades")
            print(f"✓ Price data: {price_df.index[0].strftime('%Y-%m-%d')} to {price_df.index[-1].strftime('%Y-%m-%d')} ({len(price_df)} days)")
            
            # Run live simulation
            trades, state = simulate_live_trading(insider_trades, price_df, ticker)
            
            print(f"✓ Simulation complete: {len(trades)} trades executed")
            
            # Generate event files (only for stocks with trades)
            if trades:
                generate_event_files(state.all_events, price_df, ticker)
                generate_volatility_json(state.all_events, price_df, ticker)
            
        except Exception as e:
            # Silently skip errors to keep processing
            continue
        
        if not trades:
            print(f"\n❌ No trades executed for {ticker}.")
            all_results.append({
                'ticker': ticker,
                'trade_count': 0,
                'winning_trades': 0,
                'losing_trades': 0,
                'roi': 0,
                'win_rate': 0,
                'target_rate': 0,
                'total_profit': 0,
                'total_invested': 0,
                'avg_return': 0,
                'median_return': 0,
                'avg_win': 0,
                'avg_loss': 0,
                'best_trade': 0,
                'worst_trade': 0,
                'avg_days_held': 0,
                'avg_peak_gain': 0,
                'trades': []
            })
            continue
        
        # Calculate statistics
        print()
        print("=" * 80)
        print(f"BACKTEST RESULTS - {ticker}")
        print("=" * 80)
        print()
        
        total_trades = len(trades)
        winning_trades = [t for t in trades if t['return_pct'] > 0]
        losing_trades = [t for t in trades if t['return_pct'] <= 0]
        target_reached_trades = [t for t in trades if t['target_reached'] == 'yes']
        
        win_rate = len(winning_trades) / total_trades * 100 if total_trades > 0 else 0
        target_rate = len(target_reached_trades) / total_trades * 100 if total_trades > 0 else 0
        
        total_profit = sum(t['profit_loss'] for t in trades)
        total_invested = sum(t['position_size'] for t in trades)
        roi = (total_profit / total_invested * 100) if total_invested > 0 else 0
        
        avg_return = sum(t['return_pct'] for t in trades) / total_trades
        avg_win = sum(t['return_pct'] for t in winning_trades) / len(winning_trades) if winning_trades else 0
        avg_loss = sum(t['return_pct'] for t in losing_trades) / len(losing_trades) if losing_trades else 0
        
        median_return = sorted([t['return_pct'] for t in trades])[len(trades) // 2]
        max_return = max(t['return_pct'] for t in trades)
        min_return = min(t['return_pct'] for t in trades)
        
        avg_days = sum(t['days_held'] for t in trades) / total_trades
        avg_peak_gain = sum(t['peak_gain'] for t in trades) / total_trades
        
        print(f"Total Trades:              {total_trades}")
        print(f"Winning Trades:            {len(winning_trades)} ({win_rate:.1f}%)")
        print(f"Losing Trades:             {len(losing_trades)} ({100-win_rate:.1f}%)")
        print(f"Target Reached:            {len(target_reached_trades)} ({target_rate:.1f}%)")
        print()
        print(f"Total Profit:              ${total_profit:,.2f}")
        print(f"Total Invested:            ${total_invested:,.2f}")
        print(f"ROI:                       {roi:+.2f}%")
        print()
        print(f"Average Return:            {avg_return:+.2f}%")
        print(f"Median Return:             {median_return:+.2f}%")
        print(f"Average Win:               {avg_win:+.2f}%")
        print(f"Average Loss:              {avg_loss:+.2f}%")
        print(f"Best Trade:                {max_return:+.2f}%")
        print(f"Worst Trade:               {min_return:+.2f}%")
        print()
        print(f"Average Days Held:         {avg_days:.1f}")
        print(f"Average Peak Gain:         {avg_peak_gain:.1f}%")
        print()
        
        # Exit reason breakdown
        print("EXIT REASONS:")
        print("-" * 80)
        exit_reasons = {}
        for trade in trades:
            reason = trade['sell_reason']
            if reason not in exit_reasons:
                exit_reasons[reason] = []
            exit_reasons[reason].append(trade)
        
        for reason, reason_trades in sorted(exit_reasons.items(), key=lambda x: len(x[1]), reverse=True):
            count = len(reason_trades)
            avg_ret = sum(t['return_pct'] for t in reason_trades) / count
            print(f"{reason:35} | Count: {count:2} | Avg Return: {avg_ret:+6.2f}%")
        print()
        
        # Save results in POC format for UI display
        output = {
            'ticker': ticker,
            'strategy': 'Insider Conviction (No Hindsight)',
            'total_trades': total_trades,
            'total_profit': round(total_profit, 2),
            'total_invested': total_invested,
            'roi': round(roi, 2),
            'win_rate': round(win_rate, 1),
            'target_rate': round(target_rate, 1),
            'trades': trades
        }
        
        # Save to POC file for UI
        poc_file = f'output CSVs/{ticker.lower()}_insider_conviction_poc.json'
        with open(poc_file, 'w') as f:
            json.dump(output, f, indent=2)
        print(f"✓ Results saved to: {poc_file}")
        print("=" * 80)
        
        # Add to all results
        all_results.append({
            'ticker': ticker,
            'trade_count': total_trades,  # Number of trades
            'trades': trades,  # Array of trade details (for UI)
            'winning_trades': len(winning_trades),
            'losing_trades': len(losing_trades),
            'roi': round(roi, 2),
            'win_rate': round(win_rate, 1),
            'target_rate': round(target_rate, 1),
            'total_profit': round(total_profit, 2),
            'total_invested': total_invested,
            'avg_return': round(avg_return, 2),
            'median_return': round(median_return, 2),
            'avg_win': round(avg_win, 2),
            'avg_loss': round(avg_loss, 2),
            'best_trade': round(max_return, 2),
            'worst_trade': round(min_return, 2),
            'avg_days_held': round(avg_days, 1),
            'avg_peak_gain': round(avg_peak_gain, 1)
        })
    
    # Print summary of all tickers
    print("\n\n" + "=" * 80)
    print("SUMMARY - ALL TICKERS")
    print("=" * 80)
    
    # Calculate overall statistics
    total_profit_all = sum(r['total_profit'] for r in all_results)
    total_invested_all = sum(r.get('total_invested', 0) for r in all_results)
    overall_roi = (total_profit_all / total_invested_all * 100) if total_invested_all > 0 else 0
    total_trades_all = sum(r['trade_count'] for r in all_results)
    total_winners = sum(r.get('winning_trades', 0) for r in all_results)
    overall_win_rate = (total_winners / total_trades_all * 100) if total_trades_all > 0 else 0
    
    print(f"OVERALL STATISTICS")
    print("-" * 80)
    print(f"Stocks Analyzed:           {len(all_results)}")
    print(f"Total Trades:              {total_trades_all}")
    print(f"Winning Trades:            {total_winners} ({overall_win_rate:.1f}%)")
    print(f"Total Profit:              ${total_profit_all:,.2f}")
    print(f"Total Invested:            ${total_invested_all:,.2f}")
    print(f"Overall ROI:               {overall_roi:+.2f}%")
    print()
    
    # Sort by ROI
    results_sorted = sorted(all_results, key=lambda x: x['roi'], reverse=True)
    
    # Get top 25 best and worst
    top_25_best = results_sorted[:25]
    top_25_worst = results_sorted[-25:]
    
    print("TOP 25 BEST PERFORMERS:")
    print("-" * 80)
    for result in top_25_best:
        print(f"{result['ticker']:6} | Trades: {result['trade_count']:3} | ROI: {result['roi']:+7.2f}% | Win Rate: {result['win_rate']:5.1f}% | P/L: ${result.get('total_profit', 0):,.2f}")
    
    print()
    print("TOP 25 WORST PERFORMERS:")
    print("-" * 80)
    for result in reversed(top_25_worst):
        print(f"{result['ticker']:6} | Trades: {result['trade_count']:3} | ROI: {result['roi']:+7.2f}% | Win Rate: {result['win_rate']:5.1f}% | P/L: ${result.get('total_profit', 0):,.2f}")
    
    # Get tickers of top/worst performers (for filtering detailed trades)
    top_worst_tickers = set([r['ticker'] for r in top_25_best] + [r['ticker'] for r in top_25_worst])
    
    # Remove individual trades from all_results (except top/worst performers)
    # This saves file size - we only need detailed trades for stocks shown in UI
    all_results_lite = []
    for r in results_sorted:
        result_copy = r.copy()
        if result_copy['ticker'] not in top_worst_tickers:
            # Remove trades array for stocks not in top/worst
            result_copy.pop('trades', None)
        all_results_lite.append(result_copy)
    
    # Save results to JSON for UI
    output = {
        'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'strategy': 'Insider Conviction (No Hindsight)',
        'overall_stats': {
            'stocks_analyzed': len(all_results),
            'total_trades': total_trades_all,
            'winning_trades': total_winners,
            'overall_win_rate': round(overall_win_rate, 1),
            'total_profit': round(total_profit_all, 2),
            'total_invested': total_invested_all,
            'overall_roi': round(overall_roi, 2)
        },
        'top_25_best': top_25_best,
        'top_25_worst': list(reversed(top_25_worst)),
        'all_results': all_results_lite
    }
    
    output_file = 'output CSVs/insider_conviction_all_stocks_results.json'
    with open(output_file, 'w') as f:
        json.dump(output, f, indent=2)
    
    print(f"\n✓ Results saved to: {output_file}")
    print("=" * 80)


if __name__ == "__main__":
    main()
