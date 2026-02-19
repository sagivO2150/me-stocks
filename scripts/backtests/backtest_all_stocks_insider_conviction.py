#!/usr/bin/env python3
"""
Run Insider Conviction Strategy on ALL stocks in the database.
This script generates a summary of best/worst performers for the webapp.

For detailed analysis (CSV/XLSX/JSON files), use the on-demand script:
  generate_stock_detailed_analysis.py

Output: JSON file with top 25 best/worst performers + overall ROI

NOTE: Uses cached yfinance data for speed (10-15 seconds vs hours)

USAGE:
  # Run on all stocks:
  .venv/bin/python scripts/backtests/backtest_all_stocks_insider_conviction.py
  
  # Run on single stock for testing:
  .venv/bin/python scripts/backtests/backtest_all_stocks_insider_conviction.py --ticker BSFC
"""

import pandas as pd
from datetime import datetime, timedelta
import json
from typing import List, Dict, Optional, Tuple
from enum import Enum
import argparse
import sys


class MarketPhase(Enum):
    """Current market phase we're observing."""
    RISING = "rising"
    FALLING = "falling"
    UNKNOWN = "unknown"


def load_cache_data():
    """Load cached yfinance data from JSON file"""
    cache_file = 'output CSVs/yfinance_cache_full.json'
    
    print("📦 Loading cached price data...")
    
    with open(cache_file, 'r') as f:
        cache = json.load(f)
    
    print(f"   ✅ Loaded cache with {len(cache['data'])} stocks")
    print(f"   📅 Cache created: {cache['metadata']['created']}")
    
    # Convert cache data to pandas DataFrames
    price_cache = {}
    
    for ticker, ticker_data in cache['data'].items():
        df = pd.DataFrame({
            'Open': ticker_data['open'],
            'High': ticker_data['high'],
            'Low': ticker_data['low'],
            'Close': ticker_data['close'],
            'Volume': ticker_data['volume']
        }, index=pd.to_datetime(ticker_data['dates']))
        
        price_cache[ticker] = df
    
    return price_cache


class MarketPhase(Enum):
    """Current market phase we're observing."""
    RISING = "rising"
    FALLING = "falling"
    UNKNOWN = "unknown"


class TradingState:
    """Track the current state of our live trading simulation."""
    
    def __init__(self):
        # Market observation
        self.phase = MarketPhase.UNKNOWN
        self.last_peak_date = None
        
        # Price history for rise start detection
        self.price_history = []
        
        # Dip-recovery-dip tracking for fall detection
        self.first_dip_date = None
        self.first_dip_price = None
        self.dip_low_price = None
        self.in_recovery = False
        self.recovery_high = None
        self.consecutive_up_days = 0
        
        # Current trend tracking
        self.trend_start_date = None
        self.trend_start_price = None
        self.trend_peak_price = None
        self.trend_peak_date = None
        self.trend_low_price = float('inf')
        self.trend_low_date = None
        
        # Previous trend memory
        self.prev_fall_pct = 0
        self.prev_rise_pct = 0
        self.prev_rise_start_price = None
        self.prev_rise_peak_price = None
        self.prev_fall_start_price = None
        self.prev_fall_had_insiders = False  # Track if previous fall had insider support
        
        # Insider activity tracking
        self.insiders_bought_in_rise = []
        self.insiders_bought_in_fall = []
        self.shopping_spree_peak_price = None
        
        # Position tracking
        self.in_position = False
        self.entry_date = None
        self.entry_price = None
        self.target_price = None
        self.buy_type = None
        self.position_size = 2000
        self.max_mid_fall_before_target = 0
        self.target_reached = False
        self.peak_since_entry = 0
        
        # Absorption buy specific tracking
        self.rise_start_price = None
        self.cumulative_mid_rises_pct = 0  # Sum of all mid-rise percentages in current rise event
        self.mid_rise_start_price = None  # Track start of current mid-rise
        self.in_mid_rise = False
        
        # Event tracking
        self.all_events = []  # List of all rise/fall events
    
    def update_phase(self, current_price: float, prev_price: float, date: datetime):
        """Update market phase using FIRST DIP → RECOVERY → SECOND DIP pattern for fall detection."""
        # Track price history
        self.price_history.append((date, current_price))
        if len(self.price_history) > 4:
            self.price_history.pop(0)
        
        # Treat ±$0.01 as plateau
        price_diff = abs(current_price - prev_price)
        is_plateau = price_diff <= 0.01
        is_up = current_price > prev_price and not is_plateau
        is_down = current_price < prev_price and not is_plateau
        
        # Track consecutive up days
        if is_up:
            self.consecutive_up_days += 1
        else:
            self.consecutive_up_days = 0
        
        if self.phase == MarketPhase.UNKNOWN or self.phase == MarketPhase.FALLING:
            # HUNTING FOR RISE START: Need 2 consecutive up days (or 3 if insiders bought recently)
            # Check BOTH lists in case phase detection was wrong and insiders went to wrong list
            has_insider_support = bool(self.insiders_bought_in_fall or self.insiders_bought_in_rise)
            required_up_days = 3 if has_insider_support else 2
            
            # DEBUG
            if hasattr(date, 'year') and date.year == 2025 and date.month == 4 and date.day <= 5:
                print(f"  [{date.strftime('%Y-%m-%d')}] consecutive_up={self.consecutive_up_days}, required={required_up_days}, fall_insiders={len(self.insiders_bought_in_fall)}, rise_insiders={len(self.insiders_bought_in_rise)}")
            
            if self.consecutive_up_days >= required_up_days:
                days_since_last_peak = 999  # Temporarily disable 15-day gap
                
                if days_since_last_peak >= 0:
                    
                    # Find the actual rise start
                    lookback = 4 if has_insider_support else 3
                    
                    if len(self.price_history) >= lookback:
                        bottom_date, bottom_price = self.price_history[-lookback]
                        actual_start_date = bottom_date
                        actual_start_price = bottom_price
                    else:
                        actual_start_date = date
                        actual_start_price = prev_price
                    
                    # Record the completed FALL event if we were falling before
                    if self.phase == MarketPhase.FALLING and self.trend_start_date:
                        fall_days = (actual_start_date - self.trend_start_date).days
                        fall_pct = ((self.trend_low_price - self.trend_start_price) / self.trend_start_price) * 100
                        
                        if not self.in_position:
                            self.prev_fall_pct = abs(fall_pct)
                            self.prev_fall_start_price = self.trend_start_price
                        
                        # Record fall event
                        fall_start_naive = self.trend_start_date.tz_localize(None) if hasattr(self.trend_start_date, 'tz_localize') else self.trend_start_date
                        fall_end_naive = actual_start_date.tz_localize(None) if hasattr(actual_start_date, 'tz_localize') else actual_start_date
                        
                        fall_insiders = [
                            i for i in self.insiders_bought_in_fall 
                            if fall_start_naive <= pd.to_datetime(i['date']).tz_localize(None) <= fall_end_naive
                        ]
                        
                        # Remember if this fall had insider support
                        self.prev_fall_had_insiders = len(fall_insiders) > 0
                        
                        self.all_events.append({
                            'event_type': 'DOWN',
                            'start_date': self.trend_start_date,
                            'end_date': actual_start_date,
                            'days': fall_days,
                            'change_pct': fall_pct,
                            'start_price': self.trend_start_price,
                            'end_price': self.trend_low_price,
                            'insiders': fall_insiders
                        })
                    
                    # Start rise
                    self.phase = MarketPhase.RISING
                    self.trend_start_date = actual_start_date
                    self.trend_start_price = actual_start_price
                    self.trend_peak_price = current_price
                    self.trend_peak_date = date
                    self.first_dip_date = None
                    self.first_dip_price = None
                    self.in_recovery = False
        
        elif self.phase == MarketPhase.RISING:
            # Update peak if new high
            if current_price > self.trend_peak_price:
                self.trend_peak_price = current_price
                self.trend_peak_date = date
                self.first_dip_date = None
                self.first_dip_price = None
                self.dip_low_price = None
                self.in_recovery = False
            
            # FALL DETECTION: First dip → recovery → second dip
            if is_down:
                decline_from_peak = ((self.trend_peak_price - current_price) / self.trend_peak_price) * 100
                
                if self.first_dip_date is None:
                    if decline_from_peak >= 1.0:
                        self.first_dip_date = date
                        self.first_dip_price = current_price
                        self.dip_low_price = current_price
                        self.in_recovery = False
                elif not self.in_recovery:
                    if current_price < self.dip_low_price:
                        self.dip_low_price = current_price
                elif self.in_recovery:
                    decline_from_recovery = ((self.recovery_high - current_price) / self.recovery_high) * 100
                    if decline_from_recovery >= 1.0:
                        self.prev_rise_pct = ((self.trend_peak_price - self.trend_start_price) / 
                                             self.trend_start_price) * 100
                        self.prev_rise_start_price = self.trend_start_price
                        self.prev_rise_peak_price = self.trend_peak_price
                        self.last_peak_date = self.trend_peak_date
                        
                        # Transition to FALLING
                        self.phase = MarketPhase.FALLING
                        
                        from pandas.tseries.offsets import BDay
                        actual_rise_end = self.first_dip_date - BDay(1)
                        
                        # Record completed RISE event
                        rise_days = (actual_rise_end - self.trend_start_date).days
                        rise_pct = ((self.trend_peak_price - self.trend_start_price) / self.trend_start_price) * 100
                        
                        rise_start_naive = self.trend_start_date.tz_localize(None) if hasattr(self.trend_start_date, 'tz_localize') else self.trend_start_date
                        rise_end_naive = actual_rise_end.tz_localize(None) if hasattr(actual_rise_end, 'tz_localize') else actual_rise_end
                        peak_naive = self.trend_peak_date.tz_localize(None) if hasattr(self.trend_peak_date, 'tz_localize') else self.trend_peak_date
                        
                        # Insiders who bought BEFORE the peak are rise insiders
                        # Insiders who bought AFTER the peak (during the dip) are fall insiders
                        rise_insiders = [
                            i for i in self.insiders_bought_in_rise
                            if pd.to_datetime(i['date']).tz_localize(None) <= peak_naive
                        ]
                        
                        insiders_after_peak = [
                            i for i in self.insiders_bought_in_rise
                            if pd.to_datetime(i['date']).tz_localize(None) > peak_naive
                        ]
                        
                        # Move post-peak insiders to fall list (they bought during the dip)
                        if insiders_after_peak:
                            if any('2025-03' in i['date'] or '2025-04' in i['date'] for i in insiders_after_peak):
                                print(f"  [PHASE TRANSITION] Moving {len(insiders_after_peak)} post-peak insiders from rise to fall")
                                print(f"    Insiders: {[i['date'] for i in insiders_after_peak]}")
                            self.insiders_bought_in_fall.extend(insiders_after_peak)
                        
                        self.all_events.append({
                            'event_type': 'RISE',
                            'start_date': self.trend_start_date,
                            'end_date': actual_rise_end,
                            'days': rise_days,
                            'change_pct': rise_pct,
                            'start_price': self.trend_start_price,
                            'end_price': None,
                            'peak_price': self.trend_peak_price,
                            'insiders': rise_insiders
                        })
                        
                        # Clear the rise list (all insiders already moved to fall)
                        self.insiders_bought_in_rise = []
                        
                        self.trend_start_date = self.trend_peak_date  # Fall starts from the peak
                        self.trend_start_price = self.trend_peak_price
                        self.trend_low_price = current_price
                        self.trend_low_date = date
                        self.consecutive_up_days = 0
                        
                        self.first_dip_date = None
                        self.first_dip_price = None
                        self.dip_low_price = None
                        self.in_recovery = False
            
            elif is_up and self.in_recovery:
                self.first_dip_date = None
                self.first_dip_price = None
                self.dip_low_price = None
                self.in_recovery = False
                self.recovery_high = current_price
            
            elif (is_up or is_plateau) and self.first_dip_date is not None and not self.in_recovery:
                recovery_from_dip = ((current_price - self.dip_low_price) / self.dip_low_price) * 100
                if recovery_from_dip >= 0.0:
                    self.in_recovery = True
                    self.recovery_high = current_price
        
        # Track lowest price during fall
        if self.phase == MarketPhase.FALLING:
            if current_price < self.trend_low_price:
                self.trend_low_price = current_price
                self.trend_low_date = date
        
        # Calculate current fall percentage
        if self.phase == MarketPhase.FALLING and self.trend_start_price and not self.in_position:
            self.prev_fall_pct = ((self.trend_start_price - current_price) / self.trend_start_price) * 100
    
    def record_insider_purchase(self, date_str: str, trade_info: Dict):
        """Record an insider purchase occurring today."""
        trade_data = {
            'date': date_str,
            'price': trade_info['price'],
            'insider_name': trade_info['insider_name'],
            'value': trade_info['value'],
            'stock_price': trade_info.get('stock_price', trade_info['price'])
        }
        
        # DEBUG
        if '2025-03' in date_str or '2025-04' in date_str:
            print(f"  [INSIDER {date_str}] Added to {self.phase.value} list (fall={len(self.insiders_bought_in_fall)}, rise={len(self.insiders_bought_in_rise)})")
        
        if self.phase == MarketPhase.RISING:
            self.insiders_bought_in_rise.append(trade_data)
            if self.shopping_spree_peak_price is None or trade_data['stock_price'] > self.shopping_spree_peak_price:
                self.shopping_spree_peak_price = trade_data['stock_price']
        elif self.phase == MarketPhase.FALLING:
            self.insiders_bought_in_fall.append(trade_data)
            if self.shopping_spree_peak_price is None or trade_data['stock_price'] > self.shopping_spree_peak_price:
                self.shopping_spree_peak_price = trade_data['stock_price']
    
    def check_buy_signal(self, current_date: datetime, current_price: float) -> Optional[Dict]:
        """Check if we should buy based on current state (NO HINDSIGHT)."""
        if self.in_position:
            return None
        
        if self.phase != MarketPhase.RISING:
            return None
        
        if not self.insiders_bought_in_fall:
            return None
        
        # For hot stocks, require 3 consecutive up days
        if self.insiders_bought_in_rise and self.insiders_bought_in_fall:
            if self.consecutive_up_days < 3:
                return None
        
        # SCENARIO 1: Shopping Spree
        if self.insiders_bought_in_rise and self.insiders_bought_in_fall and self.shopping_spree_peak_price:
            target = self.shopping_spree_peak_price
            
            if target > current_price:
                self.in_position = True
                self.entry_date = current_date
                self.entry_price = current_price
                self.target_price = target
                self.buy_type = 'shopping_spree'
                self.target_reached = False
                self.max_mid_fall_before_target = 0
                self.peak_since_entry = 0
                
                all_insider_purchases = self.insiders_bought_in_rise + self.insiders_bought_in_fall
                
                # Clear insider data only when buy actually happens
                self.insiders_bought_in_rise = []
                self.insiders_bought_in_fall = []
                self.shopping_spree_peak_price = None
                
                return {
                    'buy_date': current_date,
                    'entry_price': current_price,
                    'target_price': target,
                    'buy_type': 'shopping_spree',
                    'num_insiders': len(all_insider_purchases),
                    'prev_fall_pct': self.prev_fall_pct
                }
        
        # SCENARIO 2: Absorption Buy
        elif self.insiders_bought_in_fall and not self.insiders_bought_in_rise:
            total_investment = sum(abs(t['value']) for t in self.insiders_bought_in_fall)
            
            if total_investment >= 5000:
                target_gain_pct = abs(self.prev_fall_pct)
                
                self.in_position = True
                self.entry_date = current_date
                self.entry_price = current_price
                self.target_price = current_price * (1 + target_gain_pct / 100)
                self.buy_type = 'absorption_buy'
                self.target_reached = False
                self.peak_since_entry = 0
                self.rise_start_price = self.trend_start_price if self.trend_start_price else current_price
                self.in_mid_rise = False
                
                # Clear insider data only when buy actually happens
                self.insiders_bought_in_rise = []
                self.insiders_bought_in_fall = []
                self.shopping_spree_peak_price = None
                
                return {
                    'buy_date': current_date,
                    'entry_price': current_price,
                    'target_price': self.target_price,
                    'buy_type': 'absorption_buy',
                    'num_insiders': len(self.insiders_bought_in_fall),
                    'fall_pct': self.prev_fall_pct,
                    'total_investment': total_investment,
                    'rise_start_price': self.rise_start_price
                }
        
        # Don't clear insider data - let it persist until a buy happens
        return None
    
    def check_sell_signal(self, current_date: datetime, current_price: float, 
                         prev_price: float) -> Optional[Tuple[str, float]]:
        """Check if we should sell based on current conditions (NO HINDSIGHT)."""
        if not self.in_position:
            return None
        
        # Calculate current performance
        current_gain_pct = ((current_price - self.entry_price) / self.entry_price) * 100
        self.peak_since_entry = max(self.peak_since_entry, current_gain_pct)
        
        # ABSORPTION BUY has different sell logic
        if self.buy_type == 'absorption_buy':
            if self.phase == MarketPhase.RISING:
                # Track MID-RISES within the current rise event
                # Accumulate all mid-rise percentages and check if cumulative >= fall_pct
                daily_change_pct = ((current_price - prev_price) / prev_price) * 100
                
                # DEBUG for BSFC
                is_bsfc_debug = (current_date.year == 2022 and current_date.month == 1 and current_date.day >= 19)
                
                # Start a new mid-rise when going up
                if daily_change_pct > 0:
                    if not self.in_mid_rise:
                        # Starting a new mid-rise
                        self.in_mid_rise = True
                        self.mid_rise_start_price = prev_price
                        if is_bsfc_debug:
                            print(f"  📈 {current_date.strftime('%Y-%m-%d')}: Mid-rise started at ${prev_price:.2f}")
                # End mid-rise when going down
                elif daily_change_pct < -1.0 and self.in_mid_rise:
                    # Mid-rise ended - calculate and add to cumulative
                    mid_rise_pct = ((prev_price - self.mid_rise_start_price) / self.mid_rise_start_price) * 100
                    if mid_rise_pct >= 1.0:  # Only count meaningful mid-rises
                        self.cumulative_mid_rises_pct += mid_rise_pct
                        if is_bsfc_debug:
                            print(f"  📈 {current_date.strftime('%Y-%m-%d')}: Mid-rise ended: +{mid_rise_pct:.2f}%, cumulative: {self.cumulative_mid_rises_pct:.2f}%")
                    self.in_mid_rise = False
                    self.mid_rise_start_price = None
                
                # Check if cumulative mid-rises reached target
                target_gain_pct = abs(self.prev_fall_pct)
                if not self.target_reached and self.cumulative_mid_rises_pct >= target_gain_pct:
                    self.target_reached = True
                    if is_bsfc_debug:
                        print(f"  🎯 {current_date.strftime('%Y-%m-%d')}: Target reached! Cumulative: {self.cumulative_mid_rises_pct:.2f}% >= {target_gain_pct:.2f}%")
                
                # After target reached, sell on first dip
                if self.target_reached and daily_change_pct < -1.0:
                    self.in_position = False
                    if is_bsfc_debug:
                        print(f"  💰 {current_date.strftime('%Y-%m-%d')}: SELL! Dip: {daily_change_pct:.2f}%")
                    return ('absorption_target_reached', current_price)
            
            # No stop loss for absorption buy - only sell when target reached
            return None
        
        # SHOPPING SPREE sell logic
        # Emergency stop loss: -15%
        if current_gain_pct <= -15:
            self.in_position = False
            return ('stop_loss', current_price)
        
        if not self.target_reached and current_price >= self.target_price:
            self.target_reached = True
        
        daily_change_pct = ((current_price - prev_price) / prev_price) * 100
        if not self.target_reached and daily_change_pct < -1.0:
            self.max_mid_fall_before_target = max(self.max_mid_fall_before_target, 
                                                   abs(daily_change_pct))
        
        if self.target_reached:
            if daily_change_pct < -1.0:
                self.in_position = False
                return ('target_reached_first_dip', current_price)
        
        return None


def generate_event_files(events: List[Dict], price_df: pd.DataFrame, ticker: str):
    """Generate CSV and Excel files for rise/fall events with proper formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from datetime import datetime
    
    if not events:
        print("⚠️ No events to export")
        return
    
    # Look up actual end prices for RISE events (FALL events already have end_price)
    for event in events:
        if event['end_price'] is None:
            try:
                end_date = event['end_date']
                end_price = price_df.loc[end_date]['Close']
                event['end_price'] = end_price
            except:
                # Fallback: use peak price if date not found
                event['end_price'] = event.get('peak_price', event['start_price'])
    
    # Calculate cumulative percentages and ranks
    cumulative_pct = 0
    rise_events = [e for e in events if e['event_type'] == 'RISE']
    fall_events = [e for e in events if e['event_type'] == 'DOWN']
    
    # Rank rises and falls by change_pct
    rise_events_sorted = sorted(rise_events, key=lambda x: abs(x['change_pct']), reverse=True)
    fall_events_sorted = sorted(fall_events, key=lambda x: abs(x['change_pct']), reverse=True)
    
    rise_ranks = {id(e): f"{i+1}/{len(rise_events)}" for i, e in enumerate(rise_events_sorted)}
    fall_ranks = {id(e): f"{i+1}/{len(fall_events)}" for i, e in enumerate(fall_events_sorted)}
    
    # Prepare data for export
    export_data = []
    for event in events:
        cumulative_pct += event['change_pct']
        
        # Format insider purchases
        insiders_str = ""
        if event['insiders']:
            insider_dates = sorted(set([i['date'] for i in event['insiders']]))
            insiders_str = ", ".join([datetime.strptime(d, '%Y-%m-%d').strftime('%d/%m/%Y') for d in insider_dates])
        
        # Get rank
        rank = rise_ranks.get(id(event)) if event['event_type'] == 'RISE' else fall_ranks.get(id(event))
        
        export_data.append({
            'event_type': event['event_type'],
            'start_date': event['start_date'].strftime('%d/%m/%Y'),
            'end_date': event['end_date'].strftime('%d/%m/%Y'),
            'days': event['days'],
            'change_pct': round(event['change_pct'], 2),
            'cumulative_pct': round(cumulative_pct, 2),
            'insider_purchases': insiders_str if insiders_str else None,
            'rank': rank
        })
    
    # Save to CSV
    csv_file = f'output CSVs/{ticker.lower()}_rise_events.csv'
    df = pd.DataFrame(export_data)
    df.to_csv(csv_file, index=False)
    print(f"✓ CSV saved to: {csv_file}")
    
    # Create Excel with colors
    wb = Workbook()
    ws = wb.active
    ws.title = f"{ticker} Rise-Fall Events"
    
    # Colors
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    rise_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    fall_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    # Headers
    headers = ['Event Type', 'Start Date', 'End Date', 'Days', 'Change %', 'Rank', 'Cumulative %', 'Insider Purchases']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
    
    # Data rows
    for row_data in export_data:
        ws.append([
            row_data['event_type'],
            row_data['start_date'],
            row_data['end_date'],
            row_data['days'],
            row_data['change_pct'],
            row_data['rank'],
            row_data['cumulative_pct'],
            row_data['insider_purchases']
        ])
        
        # Apply color based on event type
        fill = rise_fill if row_data['event_type'] == 'RISE' else fall_fill
        for cell in ws[ws.max_row]:
            cell.fill = fill
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 30
    
    # Save Excel
    excel_file = f'output CSVs/{ticker.lower()}_rise_events.xlsx'
    wb.save(excel_file)
    print(f"✓ Excel saved to: {excel_file}")


def analyze_rise_volatility(df: pd.DataFrame, rise_event: Dict) -> Dict:
    """
    Analyze the volatility pattern during and after a rise event.
    Tracks mid-rises, mid-falls during the rise, and post-rise behavior.
    """
    start_date = rise_event['start_date']
    end_date = rise_event['end_date']
    
    rise_df = df.loc[start_date:end_date]
    
    try:
        end_idx = df.index.get_indexer([end_date], method='nearest')[0]
    except:
        return None
        
    post_rise_end_idx = min(end_idx + 31, len(df))
    post_rise_df = df.iloc[end_idx:post_rise_end_idx]
    
    result = {
        'rise_start_date': start_date.strftime('%d/%m/%Y'),
        'rise_end_date': end_date.strftime('%d/%m/%Y'),
        'rise_days': rise_event['days'],
        'rise_percentage': rise_event['change_pct'],
        'mid_rises': {},
        'mid_falls': {},
        'first_dip': None,
        'first_recovery': None,
        'second_dip': None,
        'insiders': rise_event.get('insiders', [])
    }
    
    # Track consecutive movements during the rise
    if len(rise_df) > 1:
        i = 0
        while i < len(rise_df) - 1:
            movement_start_price = rise_df['Close'].iloc[i]
            current_direction = None
            
            j = i + 1
            while j < len(rise_df):
                prev_price = rise_df['Close'].iloc[j-1]
                current_price = rise_df['Close'].iloc[j]
                
                if current_price > prev_price:
                    day_direction = 'up'
                elif current_price < prev_price:
                    day_direction = 'down'
                else:
                    j += 1
                    continue
                
                if current_direction is None:
                    current_direction = day_direction
                    j += 1
                    continue
                
                if day_direction != current_direction:
                    break
                    
                j += 1
            
            if current_direction is not None and j > i + 1:
                movement_end_price = rise_df['Close'].iloc[j-1]
                movement_end_date = rise_df.index[j-1]
                total_change_pct = ((movement_end_price - movement_start_price) / movement_start_price) * 100
                
                if current_direction == 'up' and total_change_pct >= 1.0:
                    pct_key = str(round(total_change_pct, 2))
                    result['mid_rises'][pct_key] = {
                        'date': movement_end_date.strftime('%d/%m/%Y')
                    }
                elif current_direction == 'down' and total_change_pct <= -1.0:
                    pct_key = str(round(total_change_pct, 2))
                    result['mid_falls'][pct_key] = {
                        'date': movement_end_date.strftime('%d/%m/%Y')
                    }
                
                i = j - 1
            else:
                i += 1
    
    # NEW LOGIC: Check for 2 consecutive declining mid-rises
    if len(result['mid_rises']) >= 2:
        mid_rises_with_prices = []
        for pct_key, rise_info in result['mid_rises'].items():
            date_str = rise_info['date']
            rise_date = pd.to_datetime(date_str, format='%d/%m/%Y')
            try:
                price_at_date = rise_df.loc[rise_date, 'Close']
                mid_rises_with_prices.append({
                    'date': rise_date,
                    'date_str': date_str,
                    'price': price_at_date,
                    'pct': pct_key
                })
            except:
                pass
        
        mid_rises_with_prices.sort(key=lambda x: x['date'])
        
        detected_fall_start = None
        for i in range(len(mid_rises_with_prices) - 2):
            price1 = mid_rises_with_prices[i]['price']
            price2 = mid_rises_with_prices[i + 1]['price']
            price3 = mid_rises_with_prices[i + 2]['price']
            
            if price2 < price1 and price3 < price2:
                detected_fall_start = mid_rises_with_prices[i]
                print(f"  ⚠️  DETECTED FALL PATTERN: {price1:.2f} -> {price2:.2f} -> {price3:.2f}")
                print(f"  ⚠️  Reclassifying as DOWN event starting from {detected_fall_start['date_str']}")
                
                result['rise_end_date'] = detected_fall_start['date_str']
                
                try:
                    corrected_end_date = detected_fall_start['date']
                    rise_start_price = rise_df['Close'].iloc[0]
                    corrected_end_price = rise_df.loc[corrected_end_date, 'Close']
                    corrected_rise_pct = ((corrected_end_price - rise_start_price) / rise_start_price) * 100
                    result['rise_percentage'] = corrected_rise_pct
                    
                    corrected_rise_days = (corrected_end_date - start_date).days
                    result['rise_days'] = corrected_rise_days
                    
                    print(f"  ✓ Corrected: {result['rise_start_date']} to {result['rise_end_date']} = {corrected_rise_pct:.2f}% over {corrected_rise_days} days")
                except Exception as e:
                    print(f"  ✗ Error correcting rise event: {e}")
                
                break
    
    # Analyze post-rise behavior
    if len(post_rise_df) > 1:
        peak_price = post_rise_df['Close'].iloc[0]
        
        first_dip_found = False
        first_dip_low = peak_price
        first_dip_low_idx = 0
        
        first_recovery_found = False
        recovery_high = peak_price
        recovery_high_idx = 0
        
        for i in range(1, len(post_rise_df)):
            current_price = post_rise_df['Close'].iloc[i]
            current_date = post_rise_df.index[i]
            
            if not first_dip_found:
                if current_price < first_dip_low:
                    first_dip_low = current_price
                    first_dip_low_idx = i
                    fall_pct = ((peak_price - current_price) / peak_price) * 100
                    if fall_pct >= 1.0:
                        result['first_dip'] = {
                            'date': current_date.strftime('%d/%m/%Y'),
                            'percentage': round(-fall_pct, 2),
                            'days_after_peak': i
                        }
                elif current_price > first_dip_low and result['first_dip'] is not None:
                    recovery_from_low = ((current_price - first_dip_low) / first_dip_low) * 100
                    if recovery_from_low >= 1.0:
                        first_dip_found = True
                        recovery_high = current_price
                        recovery_high_idx = i
                        result['first_recovery'] = {
                            'date': current_date.strftime('%d/%m/%Y'),
                            'percentage': round(recovery_from_low, 2),
                            'days_after_peak': i
                        }
                    
            elif not first_recovery_found:
                if current_price > recovery_high:
                    recovery_high = current_price
                    recovery_high_idx = i
                    recovery_pct = ((current_price - first_dip_low) / first_dip_low) * 100
                    if recovery_pct >= 1.0:
                        result['first_recovery'] = {
                            'date': current_date.strftime('%d/%m/%Y'),
                            'percentage': round(recovery_pct, 2),
                            'days_after_peak': i
                        }
                elif current_price < recovery_high and result['first_recovery'] is not None:
                    first_recovery_found = True
                    
            else:
                second_dip_pct = ((recovery_high - current_price) / recovery_high) * 100
                if second_dip_pct >= 1.0:
                    days_since_recovery = i - recovery_high_idx
                    if result['second_dip'] is None or second_dip_pct > abs(result['second_dip']['percentage']):
                        result['second_dip'] = {
                            'date': current_date.strftime('%d/%m/%Y'),
                            'percentage': round(-second_dip_pct, 2),
                            'days_after_peak': i,
                            'days_since_recovery': days_since_recovery
                        }
    
    return result


def generate_volatility_json(events: List[Dict], df: pd.DataFrame, ticker: str):
    """Generate JSON with volatility analysis for all rise events."""
    from datetime import datetime
    
    rise_events = [e for e in events if e['event_type'] == 'RISE']
    
    volatility_analysis = {}
    for rise_event in rise_events:
        analysis = analyze_rise_volatility(df, rise_event)
        
        if analysis is None:
            continue
        
        pct_key = str(round(rise_event['change_pct'], 2))
        volatility_analysis[pct_key] = analysis
    
    json_output = {
        'ticker': ticker,
        'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'total_rise_events': len(rise_events),
        'rise_events': volatility_analysis
    }
    
    json_file = f'output CSVs/{ticker.lower()}_rise_volatility_analysis.json'
    with open(json_file, 'w') as f:
        json.dump(json_output, f, indent=2)
    
    print(f"✓ Volatility JSON saved to: {json_file}")


def process_single_stock(ticker: str, stock_data: Dict, price_cache: Dict, generate_detailed_files: bool = False) -> Optional[Dict]:
    """
    Run the insider conviction strategy on a single stock.
    Returns summary statistics only (no detailed files).
    
    Args:
        ticker: Stock ticker symbol
        stock_data: Insider trades data for this stock
        price_cache: Pre-loaded price data from cache
        generate_detailed_files: If True, return events for file generation
    """
    try:
        # Check if we have price data in cache
        if ticker not in price_cache:
            return None
        
        price_df = price_cache[ticker]
        
        if price_df.empty or len(price_df) < 30:
            return None
        
        # Get insider trades for this ticker
        insider_trades = {}
        for trade in stock_data.get('trades', []):
            trade_date = trade.get('trade_date', '')
            if trade_date:
                try:
                    price_str = str(trade.get('price', '0')).replace('$', '').replace(',', '')
                    price = float(price_str) if price_str else 0.0
                    
                    value_str = str(trade.get('value', '0')).replace('$', '').replace('+', '').replace(',', '')
                    value = float(value_str) if value_str else 0.0
                    
                    trade_info = {
                        'price': price,
                        'insider_name': trade.get('insider_name', ''),
                        'value': value,
                        'title': trade.get('title', '')
                    }
                    
                    if trade_date not in insider_trades:
                        insider_trades[trade_date] = []
                    insider_trades[trade_date].append(trade_info)
                except:
                    continue
        
        if not insider_trades:
            return None
        
        # Run simulation
        state = TradingState()
        completed_trades = []
        
        # Debug flag for BSFC
        debug_bsfc = (ticker == "BSFC")
        debug_blne_2022 = (ticker == "BLNE")
        
        if debug_bsfc:
            print(f"\n🐛 DEBUG MODE for {ticker}:")
            print(f"   Price data: {len(price_df)} days")
            print(f"   Insider trades: {len(insider_trades)} dates")
            print()
        
        if debug_blne_2022:
            print(f"\n🐛 DEBUG MODE for {ticker} - April 2022:")
            april_2022_trades = [d for d in insider_trades.keys() if '2022-04' in d]
            print(f"   Insider trade dates in April 2022: {april_2022_trades}")
            for d in april_2022_trades:
                print(f"     {d}: {insider_trades[d]}")
            print()
        
        for i in range(1, len(price_df)):
            current_date = price_df.index[i]
            current_price = price_df['Close'].iloc[i]
            prev_price = price_df['Close'].iloc[i-1]
            date_str = current_date.strftime('%Y-%m-%d')
            
            state.update_phase(current_price, prev_price, current_date)
            
            if date_str in insider_trades:
                if debug_blne_2022 and '2022-04' in date_str:
                    print(f"  🔍 Processing {date_str}: {len(insider_trades[date_str])} trades")
                
                for trade_info in insider_trades[date_str]:
                    trade_info_with_price = trade_info.copy()
                    trade_info_with_price['stock_price'] = current_price
                    state.record_insider_purchase(date_str, trade_info_with_price)
            
            if state.in_position:
                sell_signal = state.check_sell_signal(current_date, current_price, prev_price)
                if sell_signal:
                    reason, exit_price = sell_signal
                    days_held = (current_date - state.entry_date).days
                    return_pct = ((exit_price - state.entry_price) / state.entry_price) * 100
                    profit = state.position_size * (return_pct / 100)
                    
                    if debug_bsfc:
                        print(f"   🚀 SELL: {date_str} @ ${exit_price:.2f}")
                        print(f"   Return: {return_pct:+.2f}%")
                        print(f"   Reason: {reason}")
                        print()
                    
                    trade = {
                        'entry_date': state.entry_date.strftime('%Y-%m-%d'),
                        'entry_price': round(state.entry_price, 2),
                        'exit_date': current_date.strftime('%Y-%m-%d'),
                        'exit_price': round(exit_price, 2),
                        'target_price': round(state.target_price, 2),
                        'days_held': days_held,
                        'return_pct': round(return_pct, 2),
                        'position_size': state.position_size,
                        'profit_loss': round(profit, 2),
                        'sell_reason': reason,
                        'target_reached': 'yes' if state.target_reached else 'no',
                        'peak_gain': round(state.peak_since_entry, 2),
                        'buy_type': state.buy_type
                    }
                    
                    completed_trades.append(trade)
                    
                    state.insiders_bought_in_fall = []
                    state.insiders_bought_in_rise = []
                    state.shopping_spree_peak_price = None
            else:
                buy_signal = state.check_buy_signal(current_date, current_price)
                if buy_signal and debug_bsfc:
                    print(f"   📊 BUY: {date_str} @ ${current_price:.2f}")
                    print(f"   Type: {buy_signal['buy_type']}")
                    print(f"   Target: ${buy_signal['target_price']:.2f}")
                    print(f"   Fall %: {buy_signal.get('fall_pct', 0):.2f}%")
                    print()
        
        # Handle open position at end
        if state.in_position:
            if debug_bsfc:
                print(f"   ⚠️  Position still open at end!")
                print(f"   Entry: {state.entry_date.strftime('%Y-%m-%d')} @ ${state.entry_price:.2f}")
                print(f"   Final: {price_df.index[-1].strftime('%Y-%m-%d')} @ ${price_df['Close'].iloc[-1]:.2f}")
                print(f"   Target reached: {state.target_reached}")
                print()
            
            final_date = price_df.index[-1]
            final_price = price_df['Close'].iloc[-1]
            days_held = (final_date - state.entry_date).days
            return_pct = ((final_price - state.entry_price) / state.entry_price) * 100
            profit = state.position_size * (return_pct / 100)
            
            trade = {
                'entry_date': state.entry_date.strftime('%Y-%m-%d'),
                'entry_price': round(state.entry_price, 2),
                'exit_date': final_date.strftime('%Y-%m-%d'),
                'exit_price': round(final_price, 2),
                'target_price': round(state.target_price, 2),
                'days_held': days_held,
                'return_pct': round(return_pct, 2),
                'position_size': state.position_size,
                'profit_loss': round(profit, 2),
                'sell_reason': 'end_of_period',
                'target_reached': 'yes' if state.target_reached else 'no',
                'peak_gain': round(state.peak_since_entry, 2),
                'buy_type': state.buy_type
            }
            
            completed_trades.append(trade)
        
        if not completed_trades:
            return None
        
        # Calculate summary statistics
        total_trades = len(completed_trades)
        winning_trades = [t for t in completed_trades if t['return_pct'] > 0]
        target_reached_trades = [t for t in completed_trades if t['target_reached'] == 'yes']
        
        win_rate = len(winning_trades) / total_trades * 100 if total_trades > 0 else 0
        target_rate = len(target_reached_trades) / total_trades * 100 if total_trades > 0 else 0
        
        total_profit = sum(t['profit_loss'] for t in completed_trades)
        total_invested = sum(t['position_size'] for t in completed_trades)
        roi = (total_profit / total_invested * 100) if total_invested > 0 else 0
        
        avg_return = sum(t['return_pct'] for t in completed_trades) / total_trades
        median_return = sorted([t['return_pct'] for t in completed_trades])[len(completed_trades) // 2]
        max_return = max(t['return_pct'] for t in completed_trades)
        min_return = min(t['return_pct'] for t in completed_trades)
        
        avg_days = sum(t['days_held'] for t in completed_trades) / total_trades
        
        result = {
            'ticker': ticker,
            'company_name': stock_data.get('company_name', ticker),
            'total_trades': total_trades,
            'winning_trades': len(winning_trades),
            'losing_trades': total_trades - len(winning_trades),
            'win_rate': round(win_rate, 1),
            'target_rate': round(target_rate, 1),
            'total_profit': round(total_profit, 2),
            'total_invested': total_invested,
            'roi': round(roi, 2),
            'avg_return': round(avg_return, 2),
            'median_return': round(median_return, 2),
            'max_return': round(max_return, 2),
            'min_return': round(min_return, 2),
            'avg_days_held': round(avg_days, 1),
            'trades': completed_trades  # Include individual trades for charts
        }
        
        # Add events if requested (for single-ticker detailed file generation)
        if generate_detailed_files:
            result['events'] = state.all_events
            result['price_df'] = price_df
        
        return result
        
    except Exception as e:
        print(f"❌ Error processing {ticker}: {str(e)}")
        return None


def main():
    """Run the strategy on all stocks and generate summary report."""
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Run Insider Conviction Strategy backtest')
    parser.add_argument('--ticker', type=str, help='Run backtest for a single ticker (for testing)')
    args = parser.parse_args()
    
    single_ticker = args.ticker.upper() if args.ticker else None
    
    print("=" * 80)
    if single_ticker:
        print(f"🔬 SINGLE STOCK TEST MODE: {single_ticker}")
        print("Testing changes without affecting other stocks")
    else:
        print("INSIDER CONVICTION STRATEGY - ALL STOCKS ANALYSIS")
    print("=" * 80)
    print()
    
    # Load cached price data
    price_cache = load_cache_data()
    print()
    
    # Load insider trades database
    print("Loading insider trades database...")
    with open('output CSVs/expanded_insider_trades.json', 'r') as f:
        data = json.load(f)
    
    all_stocks = data.get('data', [])
    print(f"✓ Loaded {len(all_stocks)} stocks from database")
    print()
    
    # SINGLE TICKER MODE
    if single_ticker:
        # Find the ticker in the database
        stock_data = None
        for s in all_stocks:
            if s.get('ticker', '') == single_ticker:
                stock_data = s
                break
        
        if not stock_data:
            print(f"❌ Ticker {single_ticker} not found in database")
            sys.exit(1)
        
        if single_ticker not in price_cache:
            print(f"❌ No price data found for {single_ticker}")
            sys.exit(1)
        
        print(f"🔍 Testing {single_ticker}...")
        result = process_single_stock(single_ticker, stock_data, price_cache, generate_detailed_files=True)
        
        if not result:
            print(f"⚠️  {single_ticker}: No trades generated")
            sys.exit(0)
        
        print(f"✅ {single_ticker}: {result['total_trades']} trades, {result['roi']:+.2f}% ROI")
        print()
        
        # Generate detailed XLSX and JSON files
        if 'events' in result and result['events']:
            print("📝 Generating detailed files...")
            events = result['events']
            price_df = result['price_df']
            
            # Apply corrections to rise events based on declining mid-rises pattern
            corrected_events = []
            for i, event in enumerate(events):
                if event['event_type'] == 'RISE':
                    # Analyze this rise event to detect and correct declining mid-rises
                    analysis = analyze_rise_volatility(price_df, event)
                    if analysis:
                        # Update the event with corrected data
                        new_end_date = pd.to_datetime(analysis['rise_end_date'], format='%d/%m/%Y')
                        event['end_date'] = new_end_date
                        event['days'] = analysis['rise_days']
                        event['change_pct'] = analysis['rise_percentage']
                        
                        # Filter insider trades to only include those on or before the corrected end date
                        # Any filtered insiders should be moved to the next DOWN event
                        if 'insiders' in event and event['insiders']:
                            filtered_insiders = [
                                insider for insider in event['insiders']
                                if pd.to_datetime(insider['date']) > new_end_date
                            ]
                            event['insiders'] = [
                                insider for insider in event['insiders']
                                if pd.to_datetime(insider['date']) <= new_end_date
                            ]
                            
                            # Move filtered insiders to the next DOWN event (if it exists)
                            if filtered_insiders and i + 1 < len(events):
                                next_event = events[i + 1]
                                if next_event['event_type'] == 'DOWN':
                                    # Add filtered insiders to the next DOWN event
                                    if 'insiders' not in next_event:
                                        next_event['insiders'] = []
                                    next_event['insiders'].extend(filtered_insiders)
                corrected_events.append(event)
            
            generate_event_files(corrected_events, price_df, single_ticker)
            generate_volatility_json(corrected_events, price_df, single_ticker)
            print()
        
        # Remove events and price_df from result before saving to JSON
        result.pop('events', None)
        result.pop('price_df', None)
        
        # Load existing results file
        output_file = 'output CSVs/insider_conviction_all_stocks_results.json'
        try:
            with open(output_file, 'r') as f:
                existing_data = json.load(f)
        except FileNotFoundError:
            print(f"❌ Results file not found. Run full backtest first.")
            sys.exit(1)
        
        # Find and update this ticker in all_results
        updated = False
        for i, stock in enumerate(existing_data['all_results']):
            if stock['ticker'] == single_ticker:
                existing_data['all_results'][i] = result
                updated = True
                break
        
        if not updated:
            existing_data['all_results'].append(result)
        
        # Update top_25_best if this ticker is in there
        for i, stock in enumerate(existing_data['top_25_best']):
            if stock['ticker'] == single_ticker:
                existing_data['top_25_best'][i] = result
                print(f"📈 Updated in top_25_best (rank {i+1})")
                break
        
        # Update top_25_worst if this ticker is in there
        for i, stock in enumerate(existing_data['top_25_worst']):
            if stock['ticker'] == single_ticker:
                existing_data['top_25_worst'][i] = result
                print(f"📉 Updated in top_25_worst (rank {i+1})")
                break
        
        # Update timestamp
        existing_data['analysis_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Save updated results
        with open(output_file, 'w') as f:
            json.dump(existing_data, f, indent=2)
        
        print(f"💾 Updated {output_file}")
        
        # Also update the top performers cache if ticker is in there
        cache_file = 'output CSVs/yfinance_cache_top_performers.json'
        try:
            with open(cache_file, 'r') as f:
                top_cache = json.load(f)
            
            if single_ticker in top_cache['data']:
                # Ticker is in top performers cache - already has latest data
                print(f"✅ {single_ticker} is in top performers cache")
                print()
                print("🔄 Restart webapp to see changes:")
                print("   ./restart_webapp_stocks.sh")
        except FileNotFoundError:
            pass
        
        print()
        print(f"📊 {single_ticker} TRADE DETAILS:")
        for trade in result['trades']:
            print(f"   Entry: {trade['entry_date']} @ ${trade['entry_price']}")
            print(f"   Exit:  {trade['exit_date']} @ ${trade['exit_price']}")
            print(f"   Return: {trade['return_pct']:+.2f}% ({trade['days_held']} days)")
            print(f"   Reason: {trade['sell_reason']}")
            print()
        
        print("=" * 80)
        return
    
    # FULL RUN MODE - Process each stock
    total_stocks = len(all_stocks)
    print(f"Processing {total_stocks} stocks...")
    results = []
    
    for i, stock_data in enumerate(all_stocks):
        ticker = stock_data.get('ticker', '')
        if not ticker:
            continue
        
        # Show progress for EVERY stock
        print(f"{i}/{total_stocks}", flush=True)
        
        result = process_single_stock(ticker, stock_data, price_cache)
        
        if result:
            results.append(result)
    
    print(f"\n✓ Completed processing {len(all_stocks)} stocks")
    print(f"✓ Found {len(results)} stocks with trades")
    print()
    print("=" * 80)
    
    if not results:
        print("❌ No results to analyze.")
        return
    
    # Calculate overall statistics
    total_profit = sum(r['total_profit'] for r in results)
    total_invested = sum(r['total_invested'] for r in results)
    overall_roi = (total_profit / total_invested * 100) if total_invested > 0 else 0
    
    total_trades_all = sum(r['total_trades'] for r in results)
    total_winners = sum(r['winning_trades'] for r in results)
    overall_win_rate = (total_winners / total_trades_all * 100) if total_trades_all > 0 else 0
    
    print(f"OVERALL STATISTICS")
    print("-" * 80)
    print(f"Stocks Analyzed:           {len(results)}")
    print(f"Total Trades:              {total_trades_all}")
    print(f"Winning Trades:            {total_winners} ({overall_win_rate:.1f}%)")
    print(f"Total Profit:              ${total_profit:,.2f}")
    print(f"Total Invested:            ${total_invested:,.2f}")
    print(f"Overall ROI:               {overall_roi:+.2f}%")
    print()
    
    # Sort by ROI
    results_sorted = sorted(results, key=lambda x: x['roi'], reverse=True)
    
    # Get top 25 best and worst
    top_25_best = results_sorted[:25]
    top_25_worst = results_sorted[-25:]
    
    # Get tickers of top/worst performers (for filtering detailed trades)
    top_worst_tickers = set([r['ticker'] for r in top_25_best] + [r['ticker'] for r in top_25_worst])
    
    print("TOP 25 BEST PERFORMERS:")
    print("-" * 80)
    print(f"{'Rank':<5} {'Ticker':<8} {'Trades':<7} {'ROI %':<10} {'Win Rate':<10} {'Avg Return':<12}")
    print("-" * 80)
    for i, r in enumerate(top_25_best, 1):
        print(f"{i:<5} {r['ticker']:<8} {r['total_trades']:<7} {r['roi']:+9.2f}% {r['win_rate']:>8.1f}% {r['avg_return']:+11.2f}%")
    print()
    
    print("TOP 25 WORST PERFORMERS:")
    print("-" * 80)
    print(f"{'Rank':<5} {'Ticker':<8} {'Trades':<7} {'ROI %':<10} {'Win Rate':<10} {'Avg Return':<12}")
    print("-" * 80)
    for i, r in enumerate(reversed(top_25_worst), 1):
        print(f"{i:<5} {r['ticker']:<8} {r['total_trades']:<7} {r['roi']:+9.2f}% {r['win_rate']:>8.1f}% {r['avg_return']:+11.2f}%")
    print()
    
    # Remove individual trades from all_results (except top/worst performers)
    # This saves file size - we only need detailed trades for stocks shown in UI
    all_results_lite = []
    for r in results_sorted:
        result_copy = r.copy()
        if result_copy['ticker'] not in top_worst_tickers:
            # Remove trades array for stocks not in top/worst
            result_copy.pop('trades', None)
        all_results_lite.append(result_copy)
    
    # Save results to JSON
    output = {
        'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'strategy': 'Insider Conviction (No Hindsight)',
        'overall_stats': {
            'stocks_analyzed': len(results),
            'total_trades': total_trades_all,
            'winning_trades': total_winners,
            'overall_win_rate': round(overall_win_rate, 1),
            'total_profit': round(total_profit, 2),
            'total_invested': total_invested,
            'overall_roi': round(overall_roi, 2)
        },
        'top_25_best': top_25_best,
        'top_25_worst': list(reversed(top_25_worst)),
        'all_results': all_results_lite
    }
    
    output_file = 'output CSVs/insider_conviction_all_stocks_results.json'
    with open(output_file, 'w') as f:
        json.dump(output, f, indent=2)
    
    print(f"✓ Results saved to: {output_file}")
    print()
    print("=" * 80)


if __name__ == "__main__":
    main()
